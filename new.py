import streamlit as st
import psycopg2
import plotly.graph_objects as go
import json
import re
from pathlib import Path
from datetime import datetime, time,timedelta
import time as tm
from fpdf import FPDF
#from calendario import exibir_calendario, exibir_formulario_ferias,minhas_ferias_marcadas,ferias_marcadas,add_evento
import tempfile
import os
import base64
import pandas as pd
import hashlib
from io import BytesIO
from zipfile import ZipFile
import pickle
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
st.set_page_config(page_title="Sistema de Ponto", layout="centered")
# 🔹 Função para carregar as configurações do banco de dados
def carregar_configuracao():
    try:
        with open('config.json', 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        st.error("Arquivo de configuração 'config.json' não encontrado.")
        return None
    except json.JSONDecodeError as e:
        st.error(f"Erro ao decodificar 'config.json': {e}")
        return None

# 🔹 Carregar configurações do banco de dados
DB_CONFIG = carregar_configuracao()

# 🔹 Função para obter conexão persistente
@st.cache_resource
def obter_conexao_persistente():
    if not DB_CONFIG:
        st.error("Configurações do banco não carregadas.")
        return None
    
    try:
        # 🔹 Criar conexão ao Postgre usando a sintaxe correta
        connection = psycopg2.connect(dbname=DB_CONFIG["database"], user=DB_CONFIG["user"], password=DB_CONFIG["password"], host=DB_CONFIG["host"], port=5432)
        return connection
    except KeyError as e:
        st.error(f"Configuração ausente no 'config.json': {e}")
        return None
    except Exception as e:
        st.error(f"Erro ao conectar ao banco de dados: {e}")
        return None

# 🔹 Criar a conexão persistente
conexao_persistente = obter_conexao_persistente()


def carregar_configuracoes_empresa():
    caminho_arquivo = Path("empresa.json")
    if not caminho_arquivo.exists():
        st.error("O arquivo 'empresa.json' não foi encontrado.")
        return {"almoco_minimo": "01:00:00", "limite_horas_normais": "08:40:00"}
    
    try:
        with open(caminho_arquivo, "r", encoding="utf-8") as arquivo:
            config = json.load(arquivo)
            config.setdefault("almoco_minimo", "01:00:00")
            config.setdefault("limite_horas_normais", "08:40:00")
            return config
    except Exception as e:
        st.error(f"Erro ao carregar 'empresa.json': {e}")
        return {"almoco_minimo": "01:00:00", "limite_horas_normais": "08:40:00"}


# Função para criptografar senhas
def criptografar_senha(senha):
    return hashlib.sha256(senha.encode()).hexdigest()


def tela_login():
    logo = "./logo-ligth.png"
    with open(logo, "rb") as image_file:
        base64_image = base64.b64encode(image_file.read()).decode()

    st.markdown(
        f"""
        <div style="display: flex; justify-content: center;margin-top: -20px;">
            <img src="data:image/png;base64,{base64_image}" style="width: 200px;"/>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("<h1 style='text-align: center;'>Autenticação</h1>", unsafe_allow_html=True)

    with st.form(key="login_form"):
        username = st.text_input("Usuário")
        senha = st.text_input("Senha", type="password")
        submit_button = st.form_submit_button("Entrar", use_container_width=True)

        if submit_button:
            conn = conexao_persistente
            username = username.strip()
            if conn:
                with conn.cursor() as cursor:
                    cursor.execute("""
                        SELECT id, nome, cargo, administrador, cadastro, agendamento, edita_ponto, auditoria, gerenciar_permissoes, senha
                        FROM funcionarios
                        WHERE username = %s AND ATIVO = TRUE
                    """, (username,))
                    usuario = cursor.fetchone()

                    if usuario and usuario[9] == criptografar_senha(senha):  # Índice ajustado para senha
                        st.session_state["usuario"] = {
                            "id": usuario[0],
                            "nome": usuario[1],
                            "cargo": usuario[2],
                            "administrador": usuario[3] == '1',
                            "cadastro": usuario[4] == '1',
                            "agendamento": usuario[5] == '1',
                            "edita_ponto": usuario[6] == '1',
                            "auditoria": usuario[7] == '1',
                            "gerenciar_permissoes": usuario[8] == '1',  # Novo acesso
                        }
                        st.success(f"Bem-vindo, {usuario[1]}!")
                        st.rerun()
                    else:
                        st.error("Usuário ou senha incorretos, ou usuário inativo.")


# Tela inicial

def tela_inicial():
    if "usuario" not in st.session_state:
        tela_login()
    else:
        st.empty()
        st.sidebar.image("logo-dna.png", use_container_width=True)
        usuario = st.session_state["usuario"]

        if "menu_ativo" not in st.session_state:
            st.session_state.menu_ativo = "Registrar Ponto"

        def resetar_menus(exceto=None):
            if exceto != "geral":
                st.session_state.menu_geral = None
            if exceto != "admin":
                st.session_state.menu_admin = None
            if exceto != "cadastro":
                st.session_state.menu_cadastro = None
            if exceto != "agenda":
                st.session_state.menu_agenda = None
            if exceto != "edita_ponto":
                st.session_state.menu_edita_ponto = None
            if exceto != "auditoria":  # Novo submenu
                st.session_state.menu_auditoria = None

        with st.sidebar.expander("💻 Menu Geral", expanded=False):
            escolha_geral = st.radio(
                "Opções Gerais",
                ["Registrar Ponto", "Agenda", "Meus Registros", "Minhas Faltas",
                 "Minhas Horas Extras", "Minhas Férias", "Alterar Senha"],
                index=None,
                key="menu_geral",
                on_change=resetar_menus,
                args=("geral",)
            )

        escolha_cadastro = None
        if usuario.get("cadastro"):  # Verifica apenas "cadastro" para exibir o expander
            with st.sidebar.expander("📝 Cadastros", expanded=False):
                opcoes_cadastro = ["Cadastrar Funcionário", "Lista de Funcionários", "Alterar Cadastro", "Manutenção de Senha"]
                if usuario.get("gerenciar_permissoes"):  # Adiciona a opção se tiver permissão
                    opcoes_cadastro.append("Acesso dos Funcionários")
                escolha_cadastro = st.radio(
                    "Opções do Cadastro",
                    opcoes_cadastro,
                    index=None,
                    key="menu_cadastro",
                    on_change=resetar_menus,
                    args=("cadastro",)
                )

        escolha_admin = None
        if usuario.get("administrador"):
            with st.sidebar.expander("📊 Administração", expanded=False):
                escolha_admin = st.radio(
                    "Opções Administrativas",
                    ["Dashboard", "Registros do Dia", "Folha de Ponto", "Banco de Horas", "Registro de Faltas", "Férias Marcadas"],
                    index=None,
                    key="menu_admin",
                    on_change=resetar_menus,
                    args=("admin",)
                )

        escolha_agenda = None
        if usuario.get("agendamento"):
            with st.sidebar.expander("📅 Agenda", expanded=False):
                escolha_agenda = st.radio(
                    "Opções de Agenda",
                    ["Agendar Férias", "Férias Marcadas", "Agendamento"],
                    index=None,
                    key="menu_agenda",
                    on_change=resetar_menus,
                    args=("agenda",)
                )

        escolha_editar_ponto = None
        if usuario.get("edita_ponto"):
            with st.sidebar.expander("🕒 Gerenciamento de Ponto", expanded=False):
                escolha_editar_ponto = st.radio(
                    "Gerência",
                    ["Manutenção do Ponto"],
                    index=None,
                    key="menu_edita_ponto",
                    on_change=resetar_menus,
                    args=("edita_ponto",)
                )

        escolha_auditoria = None
        if usuario.get("auditoria"):  # Novo submenu vinculado ao acesso AUDITORIA
            with st.sidebar.expander("🔍 Auditoria", expanded=False):
                escolha_auditoria = st.radio(
                    "Opções de Auditoria",
                    ["Alterações de Cadastro", "Alterações de Ponto"],
                    index=None,
                    key="menu_auditoria",
                    on_change=resetar_menus,
                    args=("auditoria",)
                )

        escolha = escolha_geral or escolha_cadastro or escolha_admin or escolha_agenda or escolha_editar_ponto or escolha_auditoria
        if escolha:
            st.session_state.menu_ativo = escolha

        if st.session_state.menu_ativo == "Registrar Ponto":
            tela_funcionario()
        elif st.session_state.menu_ativo == "Alterar Senha":
            alterar_senha()
        elif st.session_state.menu_ativo == "Agenda":
            exibir_calendario()
        elif st.session_state.menu_ativo == "Meus Registros":
            tela_periodo_trabalhado()
        elif st.session_state.menu_ativo == "Minhas Faltas":
            tela_usuario_faltas()
        elif st.session_state.menu_ativo == "Minhas Horas Extras":
            tela_banco_horas()
        elif st.session_state.menu_ativo == "Minhas Férias":
            minhas_ferias_marcadas()

        elif st.session_state.menu_ativo == "Dashboard":
            tela_dashboard_admin()
        elif st.session_state.menu_ativo == "Registros do Dia":
            tela_funcionarios_ponto_dia()
        elif st.session_state.menu_ativo == "Cadastrar Funcionário":
            cadastrar_funcionarios()
        elif st.session_state.menu_ativo == "Lista de Funcionários":
            tela_listar_funcionarios()
        elif st.session_state.menu_ativo == "Alterar Cadastro":
            tela_manutencao_funcionarios()
        elif st.session_state.menu_ativo == "Acesso dos Funcionários":  # Nova opção no submenu Cadastros
            tela_gerenciar_permissoes()

        elif st.session_state.menu_ativo == "Folha de Ponto":
            tela_periodo_trabalhado_adm()
        elif st.session_state.menu_ativo == "Banco de Horas":
            tela_banco_horas_admin()
        elif st.session_state.menu_ativo == "Registro de Faltas":
            tela_admin_faltas()
        elif st.session_state.menu_ativo == "Manutenção de Senha":
            tela_alterar_senha_admin()

        elif st.session_state.menu_ativo == "Manutenção do Ponto":
            tela_registro_ponto_manual()

        elif st.session_state.menu_ativo == "Agendar Férias":
            exibir_formulario_ferias()
        elif st.session_state.menu_ativo == "Férias Marcadas":
            ferias_marcadas()
        elif st.session_state.menu_ativo == "Agendamento":
            add_evento()

        elif st.session_state.menu_ativo == "Alterações de Cadastro":  # Nova tela
            tela_auditoria_cadastros()
        elif st.session_state.menu_ativo == "Alterações de Ponto":  # Nova tela
            tela_auditoria_pontos()

        if st.sidebar.button("Sair", use_container_width=True):
            st.session_state.clear()
            st.sidebar.success("Sessão encerrada!")
            st.rerun()

@st.cache_data
def listar_usuarios():
    if not conexao_persistente:
        st.error("Conexão com o banco de dados não disponível.")
        return []
    try:
        with conexao_persistente.cursor() as cursor:
            cursor.execute("""
                SELECT ID, NOME, USERNAME, EMAIL, DTCONTRATACAO, ADMINISTRADOR, AGENDAMENTO, EDITA_PONTO 
                FROM FUNCIONARIOS 
                ORDER BY 2

            """)
            return cursor.fetchall()
    except Exception as e:
        st.error(f"Erro ao listar usuários: {e}")
        return []


def alterar_senha_usuario(usuario_id, nova_senha):
    """Atualiza a senha do usuário no banco de dados."""
    senha_criptografada = criptografar_senha(nova_senha)
    cursor = conexao_persistente.cursor()
    cursor.execute("UPDATE FUNCIONARIOS SET SENHA = %s WHERE ID = %s", (senha_criptografada, usuario_id))
    conexao_persistente.commit()
    st.success("Senha alterada com sucesso!")


def tela_alterar_senha_admin():
    st.markdown("<h1 style='text-align: center;'>Alterar Senha de Usuários</h1>", unsafe_allow_html=True)

    usuarios = listar_usuarios()

    if not usuarios:
        st.warning("Nenhum usuário cadastrado encontrado.")
        return

    # Criar um DataFrame para facilitar o uso
    df_usuarios = pd.DataFrame(usuarios, columns=["ID", "Nome", "Username", "Email", "DtContratacao", "Administrador", "Agendamento","edita_ponto"])

    # Selectbox para escolher o usuário
    nomes_usuarios = df_usuarios["Nome"].tolist()
    usuario_selecionado = st.selectbox("Selecione um usuário para alterar a senha", options=nomes_usuarios)

    # Obter o ID do usuário selecionado
    usuario_info = df_usuarios[df_usuarios["Nome"] == usuario_selecionado].iloc[0]
    id_usuario = usuario_info["ID"]

    # Formulário para alterar a senha
    with st.form(key=f"form_usuario_{id_usuario}"):
        nova_senha = st.text_input("Digite a nova senha", type="password", key=f"senha_{id_usuario}")
        confirmar_senha = st.text_input("Confirme a nova senha", type="password", key=f"confirmar_{id_usuario}")
        alterar = st.form_submit_button("Alterar Senha", use_container_width=True)

        if alterar:
            if not nova_senha or not confirmar_senha:
                st.error("Por favor, preencha os dois campos de senha.")
            elif nova_senha != confirmar_senha:
                st.error("As senhas não coincidem. Tente novamente.")
            else:
                try:
                    alterar_senha_usuario(id_usuario, nova_senha)
                    st.success(f"Senha do usuário {usuario_selecionado} alterada com sucesso!")
                except Exception as e:
                    st.error(f"Erro ao alterar a senha: {e}")



# Função para formatar horário
def formatar_horario(horario):
    """Formata o horário no formato HH:MM:SS ou retorna 'Não registrado' se nulo."""
    if horario is None:
        return ""

    if isinstance(horario, time):
        return horario.strftime("%H:%M:%S")

    if isinstance(horario, str):
        try:
            return datetime.strptime(horario, "%H:%M:%S").strftime("%H:%M:%S")
        except ValueError:
            return ""

    return "" #Removi a exibição do não registrado no frame

def obter_registros(funcionario_id, data_inicio, data_fim):
    """Obtém registros de ponto do banco de dados para o período selecionado."""
    conn = conexao_persistente  # Certifique-se de que essa função existe e retorna a conexão com o banco
    if not conn:
        st.error("Erro ao conectar ao banco de dados.")
        return None, None, None, pd.DataFrame()

    cursor = conexao_persistente.cursor()
    try:
        # Busca informações do funcionário
        cursor.execute("SELECT NOME, CARGO, DTCADASTRO FROM FUNCIONARIOS WHERE ID = %s", (funcionario_id,))
        funcionario = cursor.fetchone()

        if not funcionario:
            st.error("Funcionário não encontrado.")
            return None, None, None, pd.DataFrame()

        nome, cargo, dtcadastro = funcionario

        # Busca registros de ponto no período, garantindo que data_inicio e data_fim sejam respeitados
        cursor.execute("""
            SELECT DATA, CHEGADA, SAIDA_ALMOCO, RETORNO_ALMOCO, SAIDA 
            FROM REGISTROS 
            WHERE FUNCIONARIO_ID = %s AND DATA BETWEEN %s AND %s
            ORDER BY DATA
        """, (funcionario_id, data_inicio, data_fim))
        registros = cursor

        if not registros:
            st.warning("Nenhum registro encontrado no período selecionado.")
            return nome, cargo, dtcadastro, pd.DataFrame()

        # Formatar registros em um DataFrame
        df = pd.DataFrame(registros, columns=["Data", "Chegada", "Saída Almoço", "Retorno Almoço", "Saída"])

        # Garantir que a coluna 'Data' seja convertida para datetime e sem hora
        df["Data"] = pd.to_datetime(df["Data"]).dt.date  # Converte para apenas data (sem hora)

        # Criar a coluna "Dia da semana"
        dias_semana = {0: "Seg", 1: "Ter", 2: "Qua", 3: "Qui", 4: "Sex", 5: "Sáb", 6: "Dom"}
        df["Dia"] = df["Data"].apply(lambda x: dias_semana[x.weekday()])

        # Criar a coluna "Data e Dia" no formato "dd/mm/yyyy - DDD"
        df["Data e Dia"] = df["Data"].apply(lambda x: x.strftime('%d/%m/%Y')) + " - " + df["Dia"]

        # Garantir que a coluna "Data e Dia" seja a primeira
        df = df[["Data e Dia", "Chegada", "Saída Almoço", "Retorno Almoço", "Saída"]]

        # Aplicar a formatação de horário
        df["Chegada"] = df["Chegada"].apply(formatar_horario)
        df["Saída Almoço"] = df["Saída Almoço"].apply(formatar_horario)
        df["Retorno Almoço"] = df["Retorno Almoço"].apply(formatar_horario)
        df["Saída"] = df["Saída"].apply(formatar_horario)

        # Adicionar dias faltantes no período solicitado
        todas_datas = pd.date_range(start=data_inicio, end=data_fim, freq='D')
        datas_existentes = df["Data e Dia"].to_list()

        # Dias que faltam no DataFrame (dentro do intervalo)
        dias_faltantes = [
            data.strftime("%d/%m/%Y") + " - " + dias_semana[data.weekday()] 
            for data in todas_datas 
            if data.strftime("%d/%m/%Y") + " - " + dias_semana[data.weekday()] not in datas_existentes
        ]
        
        # Criar um DataFrame com os dias faltantes
        dias_faltantes_df = pd.DataFrame(dias_faltantes, columns=["Data e Dia"])
        dias_faltantes_df["Chegada"] = ""
        dias_faltantes_df["Saída Almoço"] = ""
        dias_faltantes_df["Retorno Almoço"] = ""
        dias_faltantes_df["Saída"] = ""

        # Concatenar os dados existentes com os faltantes
        df = pd.concat([df, dias_faltantes_df], ignore_index=True)

        # Ordenar o DataFrame pela coluna "Data e Dia"
        df = df.sort_values(by="Data e Dia").reset_index(drop=True)

        # Centralizando os dados no dataframe
        df_style = df.style.set_properties(**{'text-align': 'center'})

        # Exibindo o dataframe centralizado
        #st.dataframe(df_style, use_container_width=True)
        st.dataframe(df,hide_index=True ,use_container_width=True)

        return nome, cargo, dtcadastro, df
    except Exception as e:
        st.error(f"Ocorreu um erro ao buscar os registros: {e}")
        return None, None, None, pd.DataFrame()
    finally:
        ''
#Cadastro da empresa
def carregar_dados_empresa():
    caminho_arquivo = Path("empresa.json")
    if not caminho_arquivo.exists():
        raise FileNotFoundError("O arquivo 'empresa.json' não foi encontrado na raiz do projeto.")
    
    with open(caminho_arquivo, "r", encoding="utf-8") as arquivo:
        return json.load(arquivo)

# Função para gerar PDF
def gerar_pdf(nome, cargo, dtcadastro, df):

    # Carregar informações da empresa do arquivo JSON
    try:
        dados_empresa = carregar_dados_empresa()
    except FileNotFoundError as e:
        print(e)
        return
        
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=5)
    pdf.set_top_margin(0)
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Título
    pdf.set_font("Arial", style="B", size=12)
    pdf.cell(0, 10, "Relatório de Ponto", ln=True, align="C")

    # Informações da empresa e do funcionário lado a lado
    pdf.set_font("Arial", size=9)

    # Primeira coluna: informações da empresa (ajustada para ser mais larga)
    x_start = pdf.get_x()
    y_start = pdf.get_y()
    pdf.multi_cell(100, 5, 
                    f"Empresa: {dados_empresa['nome']}\n"
                    f"CNPJ/CPF: {dados_empresa['cnpj']}\n"
                    f"Endereço: {dados_empresa['endereco']}\n"
                    f"Cidade/UF: {dados_empresa['cidade']} - {dados_empresa['uf']}",                   
                   border=1)
    
    # Segunda coluna: informações do funcionário (texto alinhado à direita)
    pdf.set_xy(x_start + 100, y_start)  # Move para a próxima coluna
    pdf.multi_cell(90, 5, 
                   f"136 - {nome}\n"
                   f"Função: 007 - {cargo}\n"
                   f"Admissão: {dtcadastro.strftime('%d/%m/%Y') if dtcadastro else 'Não informado'}\n"
                   f"Horário: 08:00 - 11:00 - 12:00 - 17:00", 
                   border=1, align="R")
    #pdf.ln(10)

    pdf.set_font("Arial", style="B", size=9)
    pdf.cell(34, 7, "Dia", 1, 0, "C")
    pdf.cell(39, 7, "Chegada", 1, 0, "C")
    pdf.cell(39, 7, "Saída Almoço", 1, 0, "C")
    pdf.cell(39, 7, "Retorno Almoço", 1, 0, "C")
    pdf.cell(39, 7, "Saída", 1, 1, "C")

    # Dados da tabela
    pdf.set_font("Arial", size=9)
    for _, row in df.iterrows():
        # Garantir que a "Data e Dia" seja uma string válida
        data_dia = str(row["Data e Dia"]) if pd.notnull(row["Data e Dia"]) else ""

        # Preencher as células, convertendo valores para string quando necessário
        pdf.cell(34, 6, data_dia, 1, 0, "L")
        pdf.cell(39, 6, str(row["Chegada"]), 1, 0, "C")
        pdf.cell(39, 6, str(row["Saída Almoço"]), 1, 0, "C")
        pdf.cell(39, 6, str(row["Retorno Almoço"]), 1, 0, "C")
        pdf.cell(39, 6, str(row["Saída"]), 1, 1, "C")
    
    # Adicionar as assinaturas e a data
    pdf.ln(8)
    pdf.set_font("Arial", size=9)

    # Reconhecimento e Data
    pdf.cell(0, 10, "Reconheço a exatidão destas anotações.", ln=True, align="L")
    pdf.cell(30, 10, "Data: ____/____/____", 0, 0, "L")
    pdf.ln(10)

    largura_pagina = pdf.w - 2 * pdf.l_margin
    largura_assinatura = largura_pagina / 2  # Metade da largura para cada assinatura

    # Assinatura do Funcionário (centralizado na metade esquerda)
    pdf.cell(largura_assinatura, 8, "______________________________________", 0, 0, "C")

    # Assinatura do Diretor (centralizado na metade direita)
    pdf.cell(largura_assinatura, 8, "______________________________________", 0, 1, "C")
    pdf.set_y(pdf.get_y() - 2)
    # Legendas das assinaturas
    pdf.cell(largura_assinatura, 8, "Ass. Funcionário", 0, 0, "C")
    pdf.cell(largura_assinatura, 8, "Ass. do Diretor", 0, 1, "C")

    # Criar um arquivo temporário para salvar o PDF
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    pdf.output(temp_file.name)

    return temp_file.name  # Retorna o caminho do arquivo temporário




# Função principal para exibição do relatório
def tela_periodo_trabalhado():
    """Tela principal para exibição do relatório de ponto."""
    if "usuario" not in st.session_state:
        st.warning("Faça login para acessar esta área.")
        return

    usuario = st.session_state["usuario"]

    #st.title("Relatório de Ponto")
    st.markdown("<h1 style='text-align: center;'>Relatório de Ponto</h1>", unsafe_allow_html=True)
    st.write("Selecione o período desejado para exibir os registros de ponto:")

    # Seleção de período
    col1, col2 = st.columns(2)
    with col1:
        primeiro_dia_mes = datetime(datetime.now().year, datetime.now().month, 1)
        data_inicio = st.date_input("Data de início", value=primeiro_dia_mes,format="DD/MM/YYYY")
        #st.write("Data selecionada:", data_inicio.strftime("%d/%m/%Y"))
        #data_inicio = st.date_input("Data de início", value=datetime.now() - timedelta(days=30))
    with col2:
        data_fim = st.date_input("Data de fim", value=datetime.now(),format="DD/MM/YYYY")

    if data_inicio > data_fim:
        st.error("A data de início não pode ser maior que a data de fim.")
        return

    funcionario_id = usuario.get("id")
    if not funcionario_id:
        st.error("Erro ao identificar o funcionário logado.")
        return

    # Obter registros
    nome, cargo, dtcadastro, df = obter_registros(funcionario_id, data_inicio, data_fim)

    if not df.empty:
        #st.dataframe(df)

        # Gerar PDF e oferecer download
        pdf_file = gerar_pdf(nome, cargo, dtcadastro, df)
        with open(pdf_file, "rb") as f:
            st.download_button(
                label="Baixar Relatório em PDF",
                data=f,
                file_name=f"Relatorio_Ponto_{nome.replace(' ', '_')}.pdf",
                mime="application/pdf"
            )
        os.remove(pdf_file)  # Remover arquivo temporário
    else:
        st.warning("Nenhum registro encontrado no período selecionado.")

###
##
def tela_periodo_trabalhado_adm():
    """Tela principal para exibição do relatório de ponto."""
    if "usuario" not in st.session_state:
        st.warning("Faça login para acessar esta área.")
        return

    # Conexão com o banco de dados
    conn = conexao_persistente
    if not conn:
        st.error("Erro ao conectar ao banco de dados.")
        return

    cursor = conn.cursor()
    cursor.execute("ROLLBACK")

    # Título e instruções
    st.markdown("<h1 style='text-align: center;'>Relatório de Ponto</h1>", unsafe_allow_html=True)
    st.write("Selecione o período desejado para exibir os registros de ponto:")


    # Obter lista de funcionários
    cursor.execute("SELECT ID, NOME FROM FUNCIONARIOS ORDER BY NOME")
    funcionarios = cursor.fetchall()
    opcoes_funcionarios = {f"{nome}": id_func for id_func, nome in funcionarios}

    # Seleção do funcionário
    funcionario_selecionado = st.selectbox("Selecione o Funcionário:", options=opcoes_funcionarios.keys())

    if not funcionario_selecionado:
        st.warning("Selecione um funcionário para continuar.")
        return

    funcionario_id = opcoes_funcionarios[funcionario_selecionado]

    # Seleção de período
    col1, col2 = st.columns(2)
    with col1:
        primeiro_dia_mes = datetime(datetime.now().year, datetime.now().month, 1)
        data_inicio = st.date_input("Data de início", value=primeiro_dia_mes, format="DD/MM/YYYY")
    with col2:
        data_fim = st.date_input("Data de fim", value=datetime.now(), format="DD/MM/YYYY")

    if data_inicio > data_fim:
        st.error("A data de início não pode ser maior que a data de fim.")
        return

    # Obter registros do funcionário selecionado
    nome, cargo, dtcadastro, df = obter_registros(funcionario_id, data_inicio, data_fim)

    if not df.empty:
        col1, col2 = st.columns(2)

        with col1:
            # Gerar PDF do funcionário selecionado
            pdf_file = gerar_pdf(nome, cargo, dtcadastro, df)
            with open(pdf_file, "rb") as f:
                st.download_button(
                    label="Baixar Registro do Funcionario Filtrado",
                    data=f,
                    file_name=f"Relatorio_Ponto_{nome.replace(' ', '_')}.pdf",
                    mime="application/pdf",use_container_width=True
                )
            os.remove(pdf_file)  # Remover arquivo temporário

        with col2:
            if st.button("Solicitar de Todos os Funcionários",use_container_width=True):
                # Consultar registros de todos os funcionários no período selecionado
                cursor.execute("""
                    SELECT FUNC.ID, FUNC.NOME, FUNC.CARGO, FUNC.DTCADASTRO, REG.DATA, REG.CHEGADA, REG.SAIDA_ALMOCO, REG.RETORNO_ALMOCO, REG.SAIDA
                    FROM REGISTROS REG
                    JOIN FUNCIONARIOS FUNC ON REG.FUNCIONARIO_ID = FUNC.ID
                    WHERE REG.DATA BETWEEN %s AND %s
                    ORDER BY FUNC.ID, REG.DATA
                """, (data_inicio, data_fim))
                registros = cursor.fetchall()

                if registros:
                    # Agrupar os registros por funcionário
                    funcionarios_registros = {}
                    for registro in registros:
                        funcionario_id = registro[0]
                        if funcionario_id not in funcionarios_registros:
                            funcionarios_registros[funcionario_id] = {
                                "nome": registro[1],
                                "cargo": registro[2],
                                "admissao": registro[3],
                                "registros": []
                            }
                        funcionarios_registros[funcionario_id]["registros"].append(registro[4:])

                    # Criar PDFs individuais para cada funcionário e armazená-los em um arquivo ZIP
                    zip_buffer = BytesIO()
                    with ZipFile(zip_buffer, "w") as zip_file:
                        for funcionario_id, dados in funcionarios_registros.items():
                            # Criar DataFrame para os registros do funcionário
                            registros_df = pd.DataFrame(
                                dados["registros"],
                                columns=["Data", "Chegada", "Saída Almoço", "Retorno Almoço", "Saída"]
                            )

                            # Gerar um range de datas para incluir todos os dias no período
                            todas_as_datas = pd.date_range(start=data_inicio, end=data_fim)
                            datas_df = pd.DataFrame({"Data": todas_as_datas})

                            # Mesclar para garantir que todos os dias estejam presentes
                            registros_df["Data"] = pd.to_datetime(registros_df["Data"], errors='coerce')
                            registros_completos = datas_df.merge(
                                registros_df,
                                on="Data",
                                how="left"
                            )

                            # Substituir valores nulos por vazio (campo em branco)
                            registros_completos.fillna("", inplace=True)

                            # Adicionar a coluna "Data e Dia" com nomes de dias em português (abreviados)
                            registros_completos["Data e Dia"] = registros_completos["Data"].dt.strftime("%d/%m/%Y (%a)")
                            registros_completos["Data e Dia"] = registros_completos["Data e Dia"].replace({
                                "Mon": "Seg",
                                "Tue": "Ter",
                                "Wed": "Qua",
                                "Thu": "Qui",
                                "Fri": "Sex",
                                "Sat": "Sab",
                                "Sun": "Dom"
                            }, regex=True)

                            # Formatar horários para hh:mm:ss
                            for coluna in ["Chegada", "Saída Almoço", "Retorno Almoço", "Saída"]:
                                registros_completos[coluna] = registros_completos[coluna].apply(
                                    lambda x: x.strftime("%H:%M:%S") if isinstance(x, time) else x
                                )

                            # Gerar PDF para o funcionário
                            pdf_file = gerar_pdf(
                                dados["nome"],
                                dados["cargo"],
                                dados["admissao"],
                                registros_completos
                            )

                            # Adicionar o PDF ao arquivo ZIP
                            with open(pdf_file, "rb") as f:
                                zip_file.writestr(f"Relatorio_Ponto_{dados['nome'].replace(' ', '_')}.pdf", f.read())
                            os.remove(pdf_file)  # Remover arquivo temporário

                    # Fornecer o arquivo ZIP como download imediato
                    zip_buffer.seek(0)
                    st.download_button(
                        label="Baixar",
                        data=zip_buffer,
                        file_name="Relatorios_Ponto_Todos_Funcionarios.zip",
                        mime="application/zip",use_container_width=True
                    )
                else:
                    st.warning("Nenhum registro encontrado no período selecionado.")


def verificar_restricoes_ponto(cursor, usuario_id):
    """Verifica se o usuário está de férias ou já tem uma falta registrada no dia."""
    data_atual = datetime.now().date()

    # Verificar se o usuário está de férias
    cursor.execute("""
        SELECT COUNT(*) 
        FROM FERIAS 
        WHERE FUNCIONARIO_ID = %s AND %s BETWEEN DATA_INICIO AND DATA_FIM
    """, (usuario_id, data_atual))
    esta_de_ferias = cursor.fetchone()[0] > 0

    # Verificar se já há uma falta registrada para o usuário
    cursor.execute("""
        SELECT COUNT(*) 
        FROM FALTAS 
        WHERE FUNCIONARIO_ID = %s AND DATA = %s
    """, (usuario_id, data_atual))
    falta_registrada = cursor.fetchone()[0] > 0

    if esta_de_ferias:
        st.error("Você está de férias e não pode registrar o ponto.")
        return False
    elif falta_registrada:
        st.error("Você já possui uma falta registrada para hoje. Não é possível bater o ponto.")
        return False

    return True

#Tela para o Registro do ponto


def tela_funcionario():
    if "usuario" not in st.session_state:
        st.warning("Faça login para acessar esta área.")
        return

    usuario = st.session_state["usuario"]
    nome_completo = usuario['nome']
    primeiro_nome, *sobrenomes = nome_completo.split()
    primeiro_sobrenome = sobrenomes[0] if sobrenomes else ""
    st.markdown(f"<h1 style='text-align: center; font-size: 40px;'>Olá, {primeiro_nome} {primeiro_sobrenome}!</h1>", unsafe_allow_html=True)
    st.write('______________________________________________')

    # Carregar feriados com cache
    @st.cache_data
    def carregar_feriados():
        caminho_feriados = Path("feriados.json")
        if caminho_feriados.exists():
            with open(caminho_feriados, "r", encoding="utf-8") as f:
                return {datetime.strptime(data, "%Y-%m-%d").date() for data in json.load(f)}
        st.error("Arquivo 'feriados.json' não encontrado.")
        return set()

    feriados = carregar_feriados()

    def verificar_ausencias(conn, usuario_id):
        data_atual = datetime.now().date()
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT DATA
                FROM REGISTROS
                WHERE FUNCIONARIO_ID = %s AND DATA <= %s
                ORDER BY DATA DESC
                LIMIT 1
            """, (usuario_id, data_atual - timedelta(days=1)))
            ultimo_ponto = cursor.fetchone()[0] if cursor.rowcount > 0 else None

        if not ultimo_ponto:
            ultimo_ponto = data_atual - timedelta(days=1)

        dias_faltantes = []
        with conn.cursor() as cursor:
            for dia in (ultimo_ponto + timedelta(days=i) for i in range(1, (data_atual - ultimo_ponto).days)):
                if dia.weekday() >= 5 or dia in feriados or dia == data_atual:
                    continue

                cursor.execute("""
                    SELECT EXISTS(SELECT 1 FROM REGISTROS WHERE FUNCIONARIO_ID = %s AND DATA = %s) AS registro,
                           EXISTS(SELECT 1 FROM FALTAS WHERE FUNCIONARIO_ID = %s AND DATA = %s) AS falta,
                           EXISTS(SELECT 1 FROM FERIAS WHERE FUNCIONARIO_ID = %s AND %s BETWEEN DATA_INICIO AND DATA_FIM) AS ferias
                """, (usuario_id, dia, usuario_id, dia, usuario_id, dia))
                registro, falta, ferias = cursor.fetchone()

                if not (registro or falta or ferias):
                    dias_faltantes.append(dia)

        if not dias_faltantes:
            return True

        for data in sorted(dias_faltantes):
            justificativa_key = f"justificativa_{data.strftime('%Y-%m-%d')}"
            if justificativa_key not in st.session_state:
                st.session_state[justificativa_key] = False

            if not st.session_state[justificativa_key]:
                st.warning(f"Falta registrada em {data.strftime('%d/%m/%Y')}. Justifique a ausência.")
                justificativa = st.text_area(f"Justificativa para {data.strftime('%d/%m/%Y')} (mín. 15 caracteres):", key=f"falta_{data}")
                documento = st.file_uploader(f"Anexo opcional ({data.strftime('%d/%m/%Y')}):", type=["pdf"], key=f"anexo_{data}")
                if st.button("Salvar Justificativa", key=f"salvar_{data}", use_container_width=True):
                    if len(justificativa) < 15:
                        st.error("Justificativa deve ter no mínimo 15 caracteres.")
                    else:
                        with conn.cursor() as cursor:
                            documento_blob = documento.read() if documento else None
                            cursor.execute("""
                                INSERT INTO FALTAS (FUNCIONARIO_ID, DATA, HORA, JUSTIFICATIVA, DOCUMENTO)
                                VALUES (%s, %s, %s, %s, %s)
                            """, (usuario_id, data, datetime.now().time(), justificativa, documento_blob))
                            conn.commit()
                        st.session_state[justificativa_key] = True
                        st.rerun()
        return False

    def verificar_restricoes_ponto(conn, usuario_id):
        data_atual = datetime.now().date()
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT COUNT(*)
                FROM FERIAS
                WHERE FUNCIONARIO_ID = %s AND %s BETWEEN DATA_INICIO AND DATA_FIM
            """, (usuario_id, data_atual))
            esta_de_ferias = cursor.fetchone()[0] > 0

            cursor.execute("""
                SELECT COUNT(*)
                FROM FALTAS
                WHERE FUNCIONARIO_ID = %s AND DATA = %s
            """, (usuario_id, data_atual))
            falta_registrada = cursor.fetchone()[0] > 0

            if esta_de_ferias:
                st.warning("Você está de férias hoje e não pode registrar ponto.")
                return False
            if falta_registrada:
                st.warning("Você possui uma falta registrada para hoje.")
                return False
            return True

    def calcular_horas_diarias(registros, data_atual):
        if not registros or not registros[0]:
            return timedelta(0)
        chegada = datetime.combine(data_atual, registros[0])
        saida = datetime.combine(data_atual, registros[3] or datetime.now().time())
        total_horas = saida - chegada
        if registros[1] and registros[2]:
            total_horas -= datetime.combine(data_atual, registros[2]) - datetime.combine(data_atual, registros[1])
        return total_horas

    def calcular_horas_extras(conn, registros, data_atual, funcionario_id):
        if not registros or not registros[0]:
            return
        with conn.cursor() as cursor:
            chegada = datetime.combine(data_atual, registros[0])
            saida = datetime.combine(data_atual, registros[3] or datetime.now().time())
            intervalo_almoco = timedelta()
            if registros[1] and registros[2]:
                intervalo_almoco = datetime.combine(data_atual, registros[2]) - datetime.combine(data_atual, registros[1])

            total_horas = saida - chegada - intervalo_almoco
            limite_horas = timedelta(hours=8, minutes=40)
            tolerancia_saida = datetime.combine(data_atual, time(18, 10))

            if saida > tolerancia_saida and total_horas > limite_horas:
                horas_extras = total_horas - limite_horas
                horas_extras_str = f"{horas_extras.seconds // 3600:02}:{(horas_extras.seconds // 60) % 60:02}:{horas_extras.seconds % 60:02}"
                cursor.execute("""
                    UPDATE REGISTROS
                    SET HORAEXTRA = %s
                    WHERE FUNCIONARIO_ID = %s AND DATA = %s
                """, (horas_extras_str, funcionario_id, data_atual))
                conn.commit()
                st.info(f"Notificação enviada ao administrador: {usuario['nome']} registrou {horas_extras_str} de horas extras em {data_atual.strftime('%d/%m/%Y')}.")

    def registrar_ponto(tipo, conn):
        data_atual = datetime.now().date()
        hora_atual = datetime.now().time()

        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT CHEGADA, SAIDA_ALMOCO, RETORNO_ALMOCO, SAIDA
                FROM REGISTROS
                WHERE FUNCIONARIO_ID = %s AND DATA = %s
            """, (usuario["id"], data_atual))
            registros = cursor.fetchone() or (None, None, None, None)
            chegada, saida_almoco, retorno_almoco, saida = registros

            # Validação de dependências
            validacoes = {
                "SAIDA_ALMOCO": (not chegada, "Chegada não registrada!"),
                "RETORNO_ALMOCO": (not saida_almoco, "Saída do almoço não registrada!"),
                "SAIDA": (not retorno_almoco, "Chegada não registrada!"),
                "SAIDA": (not chegada, "Chegada não registrada!")
            }
            if tipo in validacoes and validacoes[tipo][0]:
                if "mensagem_registro" not in st.session_state:
                    st.session_state.mensagem_registro = {}
                st.session_state.mensagem_registro[tipo] = validacoes[tipo][1]  # Armazena a mensagem de erro diretamente
                return

            cursor.execute(f"SELECT {tipo} FROM REGISTROS WHERE FUNCIONARIO_ID = %s AND DATA = %s", (usuario["id"], data_atual))
            resultado = cursor.fetchone()
            registro_existente = resultado[0] if resultado else None

            if registro_existente:
                if "mensagem_registro" not in st.session_state:
                    st.session_state.mensagem_registro = {}
                st.session_state.mensagem_registro[tipo] = "already_registered"
            else:
                cursor.execute(f"""
                    INSERT INTO REGISTROS (FUNCIONARIO_ID, DATA, {tipo})
                    VALUES (%s, %s, %s)
                    ON CONFLICT (FUNCIONARIO_ID, DATA)
                    DO UPDATE SET {tipo} = EXCLUDED.{tipo}
                """, (usuario["id"], data_atual, hora_atual))
                conn.commit()
                if "mensagem_registro" not in st.session_state:
                    st.session_state.mensagem_registro = {}
                st.session_state.mensagem_registro[tipo] = "registered"

                if tipo == "SAIDA":
                    calcular_horas_extras(conn, registros + (hora_atual,), data_atual, usuario["id"])
                    cursor.execute("""
                        SELECT CHEGADA, SAIDA_ALMOCO, RETORNO_ALMOCO, SAIDA
                        FROM REGISTROS
                        WHERE FUNCIONARIO_ID = %s AND DATA = %s
                    """, (usuario["id"], data_atual))
                    registros = cursor.fetchone()
                    if registros[3] and hora_atual > time(18, 10):
                        @st.dialog("Justificativa para Hora Extra")
                        def dialog_justificativa():
                            st.write("Você excedeu o limite de horas diárias. Justifique as horas extras.")
                            justificativa = st.text_area("Justificativa (mín. 15 caracteres):", key="justificativa_hora_extra")
                            col1, col2 = st.columns(2)
                            with col1:
                                if st.button("Confirmar", use_container_width=True):
                                    if len(justificativa) < 15:
                                        st.error("Justificativa deve ter no mínimo 15 caracteres.")
                                    else:
                                        cursor.execute("""
                                            UPDATE REGISTROS
                                            SET JUSTIFICATIVAHORAEXTRA = %s
                                            WHERE FUNCIONARIO_ID = %s AND DATA = %s
                                        """, (justificativa, usuario["id"], data_atual))
                                        conn.commit()
                                        st.session_state.dialog_open = False
                                        st.rerun()
                            with col2:
                                if st.button("Cancelar", use_container_width=True):
                                    st.session_state.dialog_open = False
                                    st.rerun()
                        if "dialog_open" not in st.session_state or not st.session_state.dialog_open:
                            st.session_state.dialog_open = True
                            dialog_justificativa()

    conn = conexao_persistente
    if not conn:
        st.error("Erro ao conectar ao banco de dados.")
        return

    if not verificar_ausencias(conn, usuario["id"]) or not verificar_restricoes_ponto(conn, usuario["id"]):
        return

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.button("Chegada", on_click=registrar_ponto, args=("CHEGADA", conn), use_container_width=True, key="btn_chegada")
    with col2:
        st.button("Saída Almoço", on_click=registrar_ponto, args=("SAIDA_ALMOCO", conn), use_container_width=True, key="btn_saida_almoco")
    with col3:
        st.button("Retorno Almoço", on_click=registrar_ponto, args=("RETORNO_ALMOCO", conn), use_container_width=True, key="btn_retorno_almoco")
    with col4:
        st.button("Saída", on_click=registrar_ponto, args=("SAIDA", conn), use_container_width=True, key="btn_saida")

    # Área para mensagens após os botões
    mensagem_container = st.empty()

    # Exibir mensagens temporárias
    if "mensagem_registro" in st.session_state:
        tipos = ["CHEGADA", "SAIDA_ALMOCO", "RETORNO_ALMOCO", "SAIDA"]
        for tipo in tipos:
            if tipo in st.session_state.mensagem_registro:
                mensagem = st.session_state.mensagem_registro[tipo]
                if mensagem == "already_registered":
                    mensagem_container.warning(f"⚠️ {tipo.replace('_', ' ').title()} já registrado!")
                    tm.sleep(1)
                    mensagem_container.empty()
                    del st.session_state.mensagem_registro[tipo]
                elif mensagem == "registered":
                    mensagem_container.success(f"✅ {tipo.replace('_', ' ').title()} registrado!")
                    tm.sleep(1)
                    mensagem_container.empty()
                    del st.session_state.mensagem_registro[tipo]
                else:  # Mensagem de erro de validação
                    mensagem_container.warning(f"⚠️ {mensagem}")
                    tm.sleep(1)
                    mensagem_container.empty()
                    del st.session_state.mensagem_registro[tipo]

    with conn.cursor() as cursor:
        data_atual = datetime.now().date()
        cursor.execute("""
            SELECT CHEGADA, SAIDA_ALMOCO, RETORNO_ALMOCO, SAIDA
            FROM REGISTROS
            WHERE FUNCIONARIO_ID = %s AND DATA = %s
        """, (usuario["id"], data_atual))
        registros = cursor.fetchone() or (None, None, None, None)

        if any(registros[:4]):  # Verifica se há algum registro básico
            st.markdown("<h1 style='text-align: center; font-size: 40px;'>Registros de Hoje:</h1>", unsafe_allow_html=True)
            df = pd.DataFrame([registros], columns=["Chegada", "Saída Almoço", "Retorno Almoço", "Saída"])
            df = df.applymap(lambda x: x.strftime("%H:%M:%S") if isinstance(x, time) else x if x else "Não registrado")
            st.dataframe(df, hide_index=True, use_container_width=True)

            total_horas = calcular_horas_diarias(registros, data_atual)
            total_segundos = int(total_horas.total_seconds())
            horas, resto = divmod(total_segundos, 3600)
            minutos, segundos = divmod(resto, 60)
            st.markdown(f"**Total de horas trabalhadas:** {horas:02d}:{minutos:02d}:{segundos:02d}")

        # Histórico dos últimos 7 dias
        with st.expander("Histórico dos Últimos 7 Dias"):
            data_inicio = data_atual - timedelta(days=6)
            cursor.execute("""
                SELECT DATA, CHEGADA, SAIDA_ALMOCO, RETORNO_ALMOCO, SAIDA, HORAEXTRA
                FROM REGISTROS
                WHERE FUNCIONARIO_ID = %s AND DATA BETWEEN %s AND %s
                ORDER BY DATA DESC
            """, (usuario["id"], data_inicio, data_atual))
            historico = cursor.fetchall()

            if historico:
                df_historico = pd.DataFrame(historico, columns=["Data", "Chegada", "Saída Almoço", "Retorno Almoço", "Saída", "Horas Extras"])
                df_historico["Data"] = pd.to_datetime(df_historico["Data"]).dt.strftime("%d/%m/%Y")
                df_historico = df_historico.apply(lambda x: x.apply(lambda y: y.strftime("%H:%M:%S") if isinstance(y, time) else y if y else "Não registrado"), axis=1)
                st.dataframe(df_historico, hide_index=True, use_container_width=True)
            else:
                st.info("Nenhum registro nos últimos 7 dias.")



#Manutenção do ponto caso o funcionário não esqueça de registrar o ponto

def tela_registro_ponto_manual():
    st.markdown("<h1 style='text-align: center;'>Registro de Ponto Manual</h1>", unsafe_allow_html=True)

    conn = conexao_persistente
    if not conn:
        st.error("Não foi possível conectar ao banco de dados.")
        return

    # Carregar configurações da empresa
    config_empresa = carregar_configuracoes_empresa()
    almoco_minimo = datetime.strptime(config_empresa["almoco_minimo"], "%H:%M:%S").time()
    limite_horas_normais = timedelta(
        hours=int(config_empresa["limite_horas_normais"].split(":")[0]),
        minutes=int(config_empresa["limite_horas_normais"].split(":")[1]),
        seconds=int(config_empresa["limite_horas_normais"].split(":")[2])
    )

    with conn.cursor() as cursor:
        cursor.execute("SELECT ID, NOME FROM FUNCIONARIOS WHERE ATIVO = TRUE ORDER BY NOME")
        funcionarios = cursor.fetchall()
        opcoes_funcionarios = {f"{nome}": id_func for id_func, nome in funcionarios}

        funcionario_selecionado = st.selectbox("Selecione o Funcionário:", options=opcoes_funcionarios.keys())

        if not funcionario_selecionado:
            st.warning("Selecione um funcionário para continuar.")
            return

        funcionario_id = opcoes_funcionarios[funcionario_selecionado]
        data_selecionada = st.date_input("Selecione a data do registro:", datetime.now().date(), format="DD/MM/YYYY")

        cursor.execute("""
            SELECT COUNT(*) FROM FERIAS
            WHERE FUNCIONARIO_ID = %s AND %s BETWEEN DATA_INICIO AND DATA_FIM
        """, (funcionario_id, data_selecionada))
        esta_de_ferias = cursor.fetchone()[0] > 0

        cursor.execute("""
            SELECT COUNT(*) FROM FALTAS
            WHERE FUNCIONARIO_ID = %s AND DATA = %s
        """, (funcionario_id, data_selecionada))
        falta_registrada = cursor.fetchone()[0] > 0

        if esta_de_ferias:
            st.warning("Este funcionário está de férias na data selecionada e não pode registrar ponto.")
            return
        if falta_registrada:
            st.warning("Este funcionário já possui uma falta registrada para a data selecionada.")
            return

        def calcular_horas_diarias(registros, data, hora_saida_proposta=None):
            if not registros or not registros[0]:  # Verifica se há chegada registrada
                return timedelta(0)
            chegada = datetime.combine(data, registros[0])
            saida = datetime.combine(data, hora_saida_proposta if hora_saida_proposta else (registros[3] if registros[3] else datetime.now().time()))
            total_horas = saida - chegada
            
            # Considerar intervalo de almoço apenas se ambos Saída Almoço e Retorno Almoço estiverem registrados
            intervalo_almoco = timedelta(0)
            if registros[1] and registros[2]:
                intervalo_almoco = datetime.combine(data, registros[2]) - datetime.combine(data, registros[1])
                if intervalo_almoco < timedelta(hours=almoco_minimo.hour, minutes=almoco_minimo.minute, seconds=almoco_minimo.second):
                    intervalo_almoco = timedelta(hours=almoco_minimo.hour, minutes=almoco_minimo.minute, seconds=almoco_minimo.second)
            total_horas -= intervalo_almoco
            return total_horas

        def registrar_ponto_manual(tipo, cursor, conn):
            hora_base = st.time_input(f"{tipo.replace('_', ' ').title()}:", key=f"hora_{tipo}_{data_selecionada}")
            if st.button(f"{tipo.replace('_', ' ').title()}", key=f"btn_{tipo}_{data_selecionada}", use_container_width=True):
                try:
                    # Obter a data atual combinada com a hora_base
                    hora_datetime = datetime.combine(data_selecionada, hora_base)
                    # Pegar o tempo atual completo (com segundos) no momento do clique
                    hora_atual_completa = datetime.now()
                    # Combinar a hora_base com os segundos atuais
                    hora_com_segundo = hora_datetime.replace(
                        second=hora_atual_completa.second,
                        microsecond=hora_atual_completa.microsecond
                    )
                    hora_atual = hora_com_segundo.time()

                    # Buscar o valor anterior antes de registrar (exceto para SAIDA, que será tratado no diálogo)
                    if tipo.upper() != "SAIDA":
                        cursor.execute(f"""
                            SELECT {tipo.upper()}
                            FROM REGISTROS
                            WHERE FUNCIONARIO_ID = %s AND DATA = %s
                        """, (funcionario_id, data_selecionada))
                        valor_anterior = cursor.fetchone()[0] if cursor.rowcount > 0 else None

                        # Registrar o ponto (exceto SAIDA)
                        cursor.execute(f"""
                            INSERT INTO REGISTROS (FUNCIONARIO_ID, DATA, {tipo.upper()})
                            VALUES (%s, %s, %s)
                            ON CONFLICT (FUNCIONARIO_ID, DATA)
                            DO UPDATE SET {tipo.upper()} = EXCLUDED.{tipo.upper()}
                        """, (funcionario_id, data_selecionada, hora_atual))
                        conn.commit()

                        # Registrar a alteração na tabela ALTERACOES_REGISTROS_PONTO
                        usuario_logado = st.session_state["usuario"]["nome"] if "usuario" in st.session_state else "Desconhecido"
                        cursor.execute("""
                            INSERT INTO ALTERACOES_REGISTROS_PONTO (FUNCIONARIO_ID, DATA, CAMPO_ALTERADO, VALOR_ANTERIOR, VALOR_NOVO, ALTERADO_POR)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (funcionario_id, data_selecionada, tipo.upper(), valor_anterior, hora_atual, usuario_logado))
                        conn.commit()

                    # Se for "SAIDA", calcular horas extras e solicitar justificativa antes de salvar qualquer coisa
                    if tipo.upper() == "SAIDA":
                        cursor.execute("""
                            SELECT CHEGADA, SAIDA_ALMOCO, RETORNO_ALMOCO, SAIDA
                            FROM REGISTROS
                            WHERE FUNCIONARIO_ID = %s AND DATA = %s
                        """, (funcionario_id, data_selecionada))
                        registros = cursor.fetchone()
                        total_horas = calcular_horas_diarias(registros, data_selecionada, hora_atual)

                        usuario_logado = st.session_state["usuario"]["nome"] if "usuario" in st.session_state else "Desconhecido"  # Definir aqui para passar ao diálogo

                        if total_horas > limite_horas_normais:
                            horas_extras = total_horas - limite_horas_normais
                            horas_extras_str = f"{horas_extras.seconds // 3600:02}:{(horas_extras.seconds // 60) % 60:02}:{horas_extras.seconds % 60:02}"
                            def formatar_tempo(td):
                                total_segundos = int(td.total_seconds())
                                horas, resto = divmod(total_segundos, 3600)
                                minutos, segundos = divmod(resto, 60)
                                return f"{horas:02}:{minutos:02}:{segundos:02}"

                            @st.dialog("Justificativa para Hora Extra")
                            def dialog_justificativa(horas_extras_str=horas_extras_str, usuario_logado=usuario_logado):
                                st.write(f"Total de horas trabalhadas: {formatar_tempo(total_horas)}")
                                st.write(f"Horas extras calculadas: {formatar_tempo(horas_extras)}")
                                st.write("Você excedeu o limite de horas diárias. Justifique as horas extras.")
                                justificativa = st.text_area("Justificativa (mín. 15 caracteres):", key=f"justificativa_hora_extra_{data_selecionada}")
                                col1, col2 = st.columns(2)
                                with col1:
                                    if st.button("Confirmar", key=f"confirmar_justificativa_{data_selecionada}", use_container_width=True):
                                        if len(justificativa) < 15:
                                            st.error("Justificativa deve ter no mínimo 15 caracteres.")
                                        else:
                                            with conn.cursor() as new_cursor:
                                                # Registrar a saída e a hora extra apenas após confirmação
                                                new_cursor.execute("""
                                                    INSERT INTO REGISTROS (FUNCIONARIO_ID, DATA, SAIDA, HORAEXTRA, JUSTIFICATIVAHORAEXTRA)
                                                    VALUES (%s, %s, %s, %s, %s)
                                                    ON CONFLICT (FUNCIONARIO_ID, DATA)
                                                    DO UPDATE SET SAIDA = EXCLUDED.SAIDA, HORAEXTRA = EXCLUDED.HORAEXTRA, JUSTIFICATIVAHORAEXTRA = EXCLUDED.JUSTIFICATIVAHORAEXTRA
                                                """, (funcionario_id, data_selecionada, hora_atual, horas_extras_str, justificativa))
                                                # Registrar a alteração na tabela ALTERACOES_REGISTROS_PONTO
                                                valor_anterior_saida = registros[3] if registros and registros[3] else None
                                                new_cursor.execute("""
                                                    INSERT INTO ALTERACOES_REGISTROS_PONTO (FUNCIONARIO_ID, DATA, CAMPO_ALTERADO, VALOR_ANTERIOR, VALOR_NOVO, ALTERADO_POR)
                                                    VALUES (%s, %s, %s, %s, %s, %s)
                                                """, (funcionario_id, data_selecionada, "SAIDA", valor_anterior_saida, hora_atual, usuario_logado))
                                                conn.commit()
                                            st.success("Saída e hora extra registradas com sucesso!")
                                            st.rerun()
                                with col2:
                                    if st.button("Cancelar", key=f"cancelar_justificativa_{data_selecionada}", use_container_width=True):
                                        st.info("Registro de saída e hora extra cancelado.")
                                        st.rerun()

                            dialog_justificativa(horas_extras_str=horas_extras_str, usuario_logado=usuario_logado)
                        else:
                            # Se não houver horas extras, registrar apenas a saída
                            cursor.execute(f"""
                                INSERT INTO REGISTROS (FUNCIONARIO_ID, DATA, SAIDA)
                                VALUES (%s, %s, %s)
                                ON CONFLICT (FUNCIONARIO_ID, DATA)
                                DO UPDATE SET SAIDA = EXCLUDED.SAIDA
                            """, (funcionario_id, data_selecionada, hora_atual))
                            conn.commit()

                            usuario_logado = st.session_state["usuario"]["nome"] if "usuario" in st.session_state else "Desconhecido"
                            cursor.execute("""
                                INSERT INTO ALTERACOES_REGISTROS_PONTO (FUNCIONARIO_ID, DATA, CAMPO_ALTERADO, VALOR_ANTERIOR, VALOR_NOVO, ALTERADO_POR)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            """, (funcionario_id, data_selecionada, "SAIDA", registros[3] if registros else None, hora_atual, usuario_logado))
                            conn.commit()

                    placeholder = st.empty()
                    placeholder.success(f"Registrado!", icon="✅")
                    tm.sleep(1)
                    placeholder.empty()

                except Exception as e:
                    st.error(f"Erro ao registrar {tipo}: {e}")

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            registrar_ponto_manual("CHEGADA", cursor, conn)
        with col2:
            registrar_ponto_manual("SAIDA_ALMOCO", cursor, conn)
        with col3:
            registrar_ponto_manual("RETORNO_ALMOCO", cursor, conn)
        with col4:
            registrar_ponto_manual("SAIDA", cursor, conn)

        cursor.execute("""
            SELECT CHEGADA, SAIDA_ALMOCO, RETORNO_ALMOCO, SAIDA
            FROM REGISTROS
            WHERE FUNCIONARIO_ID = %s AND DATA = %s
        """, (funcionario_id, data_selecionada))
        registros = cursor.fetchone()

        if registros:
            st.markdown("<h2 style='text-align: center;'>Registros do Dia</h2>", unsafe_allow_html=True)
            df = pd.DataFrame([registros], columns=["Chegada", "Saída Almoço", "Retorno Almoço", "Saída"])
            df = df.applymap(lambda x: x.strftime("%H:%M:%S") if x else "Não registrado")
            st.dataframe(df, hide_index=True, use_container_width=True)
        else:
            st.info("Nenhum registro encontrado para a data selecionada.")

###
#Banco de horas
def tela_banco_horas():
    usuario = st.session_state["usuario"]
    st.markdown(f"<h1 style='text-align: center;'>Minhas Horas Extras</h1>", unsafe_allow_html=True)

    conn = conexao_persistente
    if conn:
        cursor = conexao_persistente.cursor()
        try:
            # Consulta ajustada para excluir valores nulos
            cursor.execute("""
                SELECT DATA, HORAEXTRA,JUSTIFICATIVAHORAEXTRA
                FROM REGISTROS
                WHERE FUNCIONARIO_ID = %s AND HORAEXTRA IS NOT NULL
                ORDER BY DATA
            """, (usuario["id"],))
            registros = cursor.fetchall()

            if registros:
                # Criando o DataFrame
                df = pd.DataFrame(registros, columns=["Data", "Hora Extra", "Justifivativa"])

                # Convertendo 'Data' para datetime e formatando como 'dd/mm/yyyy'
                df["Data"] = pd.to_datetime(df["Data"]).dt.strftime("%d/%m/%Y")

                # Convertendo 'Hora Extra' para string no formato 'HH:MM:SS'
                def format_hora_extra(time_value):
                    if isinstance(time_value, str):  # Caso já seja string no formato "HH:MM:SS"
                        return time_value
                    elif isinstance(time_value, time):  # Caso seja um objeto time
                        return time_value.strftime("%H:%M:%S")
                    else:
                        st.warning(f"Valor inesperado em Hora Extra: {time_value}")
                        return "00:00:00"

                df["Hora Extra"] = df["Hora Extra"].apply(format_hora_extra)

                # Definindo a coluna 'Data' como índice
                #df.set_index("Data", inplace=True)

                st.dataframe(df,hide_index=True,use_container_width=True)

                # Calculando o total de horas extras
                total_horas = timedelta()
                for hora in df["Hora Extra"]:
                    h, m, s = map(int, hora.split(":"))
                    total_horas += timedelta(hours=h, minutes=m, seconds=s)

                total_horas_em_horas = total_horas.total_seconds() / 3600 if total_horas else 0

                # Calculando dias adicionais (1 dia = 10 horas)
                dias_adicionais = int(total_horas_em_horas // 10)
                horas_restantes = total_horas_em_horas % 10

                # Calculando os minutos restantes após a parte das horas
                minutos_restantes = (total_horas.total_seconds() % 3600) / 60

                # Exibindo horas extras como "X Dia(s) e HH:MM"
                if total_horas_em_horas >= 10:
                    horas_extras_display = f"{dias_adicionais} Dia(s) e {int(horas_restantes):02d}:{int(minutos_restantes):02d}"
                else:
                    horas_extras_display = f"{int(horas_restantes):02d}:{int(minutos_restantes):02d}"

                # Exibindo resultados
                st.markdown(f"**Horas Extras:** {horas_extras_display}")
            else:
                st.warning("Você não tem horas extras!")
        except Exception as e:
            st.error(f"Erro ao executar a consulta: {e}")
    else:
        st.error("Não foi possível conectar ao banco de dados.")


##
##Banco de horas ADM
def tela_banco_horas_admin():
    st.markdown(f"<h1 style='text-align: center;'>Banco de Horas</h1>", unsafe_allow_html=True)

    conn = conexao_persistente
    if conn:
        cursor = conexao_persistente.cursor()
        try:
            # Checkbox para incluir funcionários inativos
            exibir_inativos = st.checkbox("Exibir Funcionários Inativos", value=False)

            # Ajustar a query para incluir ou excluir inativos
            query = """
                SELECT f.NOME AS Funcionario, r.DATA, r.HORAEXTRA, r.justificativahoraextra
                FROM REGISTROS r
                JOIN FUNCIONARIOS f ON r.FUNCIONARIO_ID = f.ID
                WHERE r.HORAEXTRA IS NOT NULL
            """
            if not exibir_inativos:
                query += " AND f.ATIVO = TRUE"
            query += " ORDER BY f.NOME, r.DATA"

            cursor.execute(query)
            registros = cursor.fetchall()

            if registros:
                # Criando o DataFrame para exibir detalhes diários
                df_detalhes = pd.DataFrame(registros, columns=["Funcionário", "Data", "Hora Extra", "Justificativa"])

                # Convertendo 'Data' para datetime e formatando como 'dd/mm/yyyy'
                df_detalhes["Data"] = pd.to_datetime(df_detalhes["Data"]).dt.strftime("%d/%m/%Y")

                # Convertendo 'Hora Extra' para string no formato 'HH:MM:SS'
                def format_hora_extra(time_value):
                    if isinstance(time_value, str):
                        return time_value
                    elif isinstance(time_value, time):
                        return time_value.strftime("%H:%M:%S")
                    else:
                        return "00:00:00"

                df_detalhes["Hora Extra"] = df_detalhes["Hora Extra"].apply(format_hora_extra)

                # Filtro para selecionar um funcionário específico
                funcionarios_unicos = df_detalhes["Funcionário"].unique()
                funcionario_selecionado = st.selectbox("Selecione um Funcionário", options=funcionarios_unicos)

                # Filtrando o DataFrame com base no funcionário selecionado
                df_filtrado = df_detalhes[df_detalhes["Funcionário"] == funcionario_selecionado]

                # Exibindo o DataFrame filtrado de detalhes
                st.subheader("Detalhes por Dia")
                st.dataframe(df_filtrado, use_container_width=True, hide_index=True)

                # Botões de download para o funcionário filtrado
                col1, col2 = st.columns(2)
                with col1:
                    # Download em XLSX (filtrado)
                    buffer = BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        df_filtrado.to_excel(writer, index=False, sheet_name='Horas_Extras_Filtrado')
                        workbook = writer.book
                        worksheet = writer.sheets['Horas_Extras_Filtrado']
                        for column_cells in worksheet.columns:
                            length = max(len(str(cell.value or "")) for cell in column_cells)
                            col_letter = get_column_letter(column_cells[0].column)
                            worksheet.column_dimensions[col_letter].width = length + 2
                            for cell in column_cells:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                    buffer.seek(0)
                    st.download_button(
                        label="Baixar Filtrado em XLSX",
                        data=buffer,
                        file_name=f"Horas_Extras_{funcionario_selecionado}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                with col2:
                    # Download em PDF (filtrado)
                    pdf = FPDF()
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf.set_top_margin(15)  # Margem superior
                    pdf.add_page()
                    pdf.set_font("Arial", "B", 12)
                    
                    # Cabeçalho
                    config_empresa = carregar_configuracoes_empresa()
                    nome_empresa = config_empresa.get("nome", "Nome da Empresa")
                    data_emissao = datetime.now().strftime("%d/%m/%Y")
                    pdf.set_font("Arial", "", 9)  # Fonte menor para data
                    pdf.cell(0, 10, f"Data de Emissão: {data_emissao}", ln=True, align="R")
                    pdf.set_font("Arial", "B", 12)
                    pdf.cell(0, 10, nome_empresa, ln=True, align="C")
                    pdf.ln(5)
                    
                    pdf.cell(0, 10, f"Horas Extras - {funcionario_selecionado}", ln=True, align="C")
                    pdf.ln(5)
                    
                    pdf.set_font("Arial", "B", 9)
                    pdf.cell(50, 7, "Funcionário", 1, 0, "C")
                    pdf.cell(30, 7, "Data", 1, 0, "C")
                    pdf.cell(30, 7, "Hora Extra", 1, 0, "C")
                    pdf.cell(80, 7, "Justificativa", 1, 1, "C")
                    pdf.set_font("Arial", "", 9)
                    for _, row in df_filtrado.iterrows():
                        pdf.cell(50, 6, row["Funcionário"], 1, 0, "C")
                        pdf.cell(30, 6, row["Data"], 1, 0, "C")
                        pdf.cell(30, 6, row["Hora Extra"], 1, 0, "C")
                        pdf.cell(80, 6, row["Justificativa"] if row["Justificativa"] else "", 1, 1, "C")
                    
                    pdf_buffer = BytesIO()
                    pdf_bytes = pdf.output(dest='S').encode('latin1')
                    pdf_buffer.write(pdf_bytes)
                    pdf_buffer.seek(0)
                    st.download_button(
                        label="Baixar Filtrado em PDF",
                        data=pdf_buffer,
                        file_name=f"Horas_Extras_{funcionario_selecionado}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )

                # Criando o DataFrame para exibir totais por funcionário
                total_horas_por_funcionario = {}
                for _, row in df_detalhes.iterrows():
                    funcionario = row["Funcionário"]
                    h, m, s = map(int, row["Hora Extra"].split(":"))
                    horas_extras = timedelta(hours=h, minutes=m, seconds=s)
                    if funcionario in total_horas_por_funcionario:
                        total_horas_por_funcionario[funcionario] += horas_extras
                    else:
                        total_horas_por_funcionario[funcionario] = horas_extras

                # Preparando o DataFrame de totais
                totais_data = []
                for funcionario, total_timedelta in total_horas_por_funcionario.items():
                    total_horas = total_timedelta.total_seconds() / 3600
                    dias_adicionais = int(total_horas // 10)
                    horas_restantes = int(total_horas % 10)
                    minutos_restantes = int((total_timedelta.total_seconds() % 3600) / 60)
                    total_formatado = f"{dias_adicionais} Dia(s) e {horas_restantes:02d}:{minutos_restantes:02d}"
                    totais_data.append([funcionario, total_formatado])

                df_totais = pd.DataFrame(totais_data, columns=["Funcionário", "Total de Horas Extras"])

                # Exibindo o DataFrame de totais
                st.subheader("Totais por Funcionário")
                st.dataframe(df_totais, use_container_width=True, hide_index=True)

                # Botões de download para totais
                col3, col4 = st.columns(2)
                with col3:
                    # Download em XLSX (totais)
                    buffer = BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        df_totais.to_excel(writer, index=False, sheet_name='Totais_Horas_Extras')
                        workbook = writer.book
                        worksheet = writer.sheets['Totais_Horas_Extras']
                        for column_cells in worksheet.columns:
                            length = max(len(str(cell.value or "")) for cell in column_cells)
                            col_letter = get_column_letter(column_cells[0].column)
                            worksheet.column_dimensions[col_letter].width = length + 2
                            for cell in column_cells:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                    buffer.seek(0)
                    st.download_button(
                        label="Baixar Totais em XLSX",
                        data=buffer,
                        file_name="Totais_Horas_Extras.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                with col4:
                    # Criar objeto PDF
                    pdf = FPDF()
                    pdf.set_auto_page_break(auto=True, margin=8)
                    pdf.set_top_margin(8)  # Margem superior
                    pdf.add_page()
                    pdf.set_font("Arial", "B", 12)

                    # Carregar informações da empresa
                    config_empresa = carregar_configuracoes_empresa()
                    nome_empresa = config_empresa.get("nome", "Nome da Empresa")
                    data_emissao = datetime.now().strftime("%d/%m/%Y")

                    # Definir coordenadas e tamanho do cabeçalho
                    x_margem = 10  # Margem lateral
                    y_margem = 10  # Margem superior do cabeçalho
                    largura = 190  # Largura do retângulo (ajustado para alinhar no A4)
                    altura = 15    # Altura do retângulo

                    # Criar retângulo ao redor do cabeçalho
                    pdf.rect(x_margem, y_margem, largura, altura)

                    # Posicionar no topo do retângulo
                    pdf.set_xy(x_margem, y_margem + 3)

                    # Nome da empresa (esquerda) e data (direita) na mesma linha
                    pdf.set_font("Arial", "B", 12)
                    pdf.cell(largura * 0.7, 5, nome_empresa, align="L")  # 70% da largura

                    pdf.set_font("Arial", "", 10)
                    pdf.cell(largura * 0.3, 5, f"Emissão: {data_emissao}", align="R")  # 30% da largura

                    pdf.ln(10)  # Espaço após cabeçalho

                    # Título do relatório
                    pdf.cell(0, 10, "Totais de Horas Extras por Funcionário", ln=True, align="C")
                    pdf.ln(5)

                    
                    pdf.set_font("Arial", "B", 9)
                    pdf.cell(100, 7, "Funcionário", 1, 0, "C")
                    pdf.cell(90, 7, "Total de Horas Extras", 1, 1, "C")
                    pdf.set_font("Arial", "", 9)
                    for _, row in df_totais.iterrows():
                        pdf.cell(100, 6, row["Funcionário"], 1, 0, "C")
                        pdf.cell(90, 6, row["Total de Horas Extras"], 1, 1, "C")
                    
                    pdf_buffer = BytesIO()
                    pdf_bytes = pdf.output(dest='S').encode('latin1')
                    pdf_buffer.write(pdf_bytes)
                    pdf_buffer.seek(0)
                    st.download_button(
                        label="Baixar Totais em PDF",
                        data=pdf_buffer,
                        file_name="Totais_Horas_Extras.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )

            else:
                st.warning("Nenhum registro de horas extras encontrado.")
        except Exception as e:
            st.error(f"Erro ao executar a consulta: {e}")
    else:
        st.error("Não foi possível conectar ao banco de dados.")

# Função para gerar o hash de uma senha
def gerar_hash_senha(senha):
    return hashlib.sha256(senha.encode()).hexdigest()

# Função para alterar a senha
def alterar_senha():
    st.markdown("<h1 style='text-align: center;'>Alteração de Senha</h1>", unsafe_allow_html=True)

    if "usuario" not in st.session_state:
        st.warning("Faça login para acessar esta área.")
        return

    usuario = st.session_state["usuario"]
    with st.form(key="alterar_senha_form"):
        senha_atual = st.text_input("Senha Atual", type="password")
        nova_senha = st.text_input("Nova Senha", type="password")
        confirmar_senha = st.text_input("Confirmar Nova Senha", type="password")
        submit_button = st.form_submit_button("Alterar Senha", use_container_width=True)

        if submit_button:
            if not senha_atual or not nova_senha or not confirmar_senha:
                st.error("Por favor, preencha todos os campos.")
                return

            if len(nova_senha) < 4:  # Verificar se a nova senha tem pelo menos 4 caracteres
                st.error("A nova senha deve ter pelo menos 4 caracteres.")
                return

            if nova_senha != confirmar_senha:
                st.error("A nova senha e a confirmação não coincidem.")
                return

            conn = conexao_persistente
            if conn:
                cursor = conexao_persistente.cursor()

                # Verificar se a senha atual está correta
                cursor.execute("""
                    SELECT senha
                    FROM funcionarios
                    WHERE id = %s
                """, (usuario["id"],))
                registro = cursor.fetchone()

                if not registro or registro[0] != gerar_hash_senha(senha_atual):
                    st.error("Senha atual incorreta.")
                else:
                    # Atualizar a senha no banco de dados
                    nova_senha_hash = gerar_hash_senha(nova_senha)
                    cursor.execute("""
                        UPDATE funcionarios
                        SET senha = %s
                        WHERE id = %s
                    """, (nova_senha_hash, usuario["id"]))
                    conn.commit()
                    st.success("Senha alterada com sucesso!")
            else:
                st.error("Erro ao conectar ao banco de dados.")

#######################################################################
#Validações do Cadastro
#######################################################################

def limpar_texto(texto, campo, tamanho_max):
    """Remove espaços extras e limita o tamanho do texto."""
    if not texto:
        return ""
    texto_limpo = " ".join(texto.split()).strip()
    if len(texto_limpo) > tamanho_max:
        st.warning(f"O campo '{campo}' foi truncado para {tamanho_max} caracteres.")
        return texto_limpo[:tamanho_max]
    return texto_limpo

def username_em_uso(username):
    """Verifica se o username já está em uso no banco."""
    conn = conexao_persistente
    with conn.cursor() as cursor:
        cursor.execute("SELECT COUNT(*) FROM FUNCIONARIOS WHERE USERNAME = %s", (username,))
        return cursor.fetchone()[0] > 0

def validar_email(email):
    """Valida o formato do e-mail com uma regex simples."""
    if not email:
        return False
    return bool(re.match(r"[^@]+@[^@]+\.[^@]+", email))



###################################################
#Função de manutenção do cadastro de funcinoários
###################################################
def tela_manutencao_funcionarios():
    st.markdown("<h1 style='text-align: center;'>Manutenção de Funcionários</h1>", unsafe_allow_html=True)

    # Forçar a limpeza do cache ao iniciar a tela
    if "cache_cleared" not in st.session_state:
        st.cache_data.clear()
        st.session_state["cache_cleared"] = True

    # Inicializar estados necessários se não existirem
    if "alterar_dados" not in st.session_state:
        st.session_state.alterar_dados = True
    if "alterar_senha" not in st.session_state:
        st.session_state.alterar_senha = False
    if "cadastro_atual" not in st.session_state:
        st.session_state.cadastro_atual = False

    conn = conexao_persistente
    if not conn:
        st.error("Erro ao conectar ao banco de dados.")
        return

    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT ID, NOME, USERNAME, EMAIL, DTCONTRATACAO, CARGO, ATIVO FROM FUNCIONARIOS ORDER BY NOME")
            usuarios = cursor.fetchall()
    except Exception as e:
        st.error(f"Erro ao listar usuários: {e}")
        return

    if not usuarios:
        st.warning("Nenhum funcionário encontrado.")
        return

    df_usuarios = pd.DataFrame(usuarios, columns=["ID", "Nome", "Username", "Email", "DtContratacao", "Cargo", "Ativo"])
    funcionarios_unicos = df_usuarios["Nome"].unique()
    funcionario_selecionado = st.selectbox("Selecione um Funcionário", options=funcionarios_unicos)

    df_filtrado = df_usuarios[df_usuarios["Nome"] == funcionario_selecionado]
    if df_filtrado.empty:
        st.warning("Funcionário não encontrado.")
        return

    funcionario = df_filtrado.iloc[0]
    id_usuario = int(funcionario["ID"])
    nome_atual = funcionario["Nome"]
    username_atual = funcionario["Username"]
    email_atual = funcionario["Email"]
    dt_contratacao_atual = funcionario["DtContratacao"]
    cargo_atual = funcionario["Cargo"] or ""
    ativo_atual = True if pd.isna(funcionario["Ativo"]) else bool(funcionario["Ativo"])

    # Atualizar valores no session_state ao trocar de usuário
    if "usuario_selecionado" not in st.session_state or st.session_state.usuario_selecionado != funcionario_selecionado:
        st.session_state.usuario_selecionado = funcionario_selecionado
        st.session_state.ativo_atual = ativo_atual

    col1, col2 = st.columns(2)
    with col1:
        alterar_dados = st.checkbox(
            "Alterar Dados Cadastrais",
            value=st.session_state.alterar_dados,
            key="alterar_dados",
            on_change=lambda: st.session_state.update({"alterar_dados": True, "alterar_senha": False})
        )
    with col2:
        alterar_senha = st.checkbox(
            "Alterar Senha",
            value=st.session_state.alterar_senha,
            key="alterar_senha",
            on_change=lambda: st.session_state.update({"alterar_senha": True, "alterar_dados": False})
        )

    if alterar_dados:
        with st.form("form_dados_cadastrais"):
            novo_nome = st.text_input("Nome do Funcionário", value=nome_atual, max_chars=100)
            novo_username = st.text_input("Nome de Usuário", value=username_atual, max_chars=50)
            novo_cargo = st.text_input("Cargo do Funcionário", value=cargo_atual, max_chars=50)
            novo_email = st.text_input("Email", value=email_atual, max_chars=100)
            nova_dt_contratacao = st.date_input(
                "Data de Contratação",
                value=dt_contratacao_atual or datetime.today().date(),
                format='DD/MM/YYYY'
            )
            novo_ativo = st.checkbox("Func. Ativo", value=st.session_state.ativo_atual, key="checkbox_ativo")

            submit_button = st.form_submit_button("Salvar Alterações", use_container_width=True)

            if submit_button:
                erros = []
                novo_nome = limpar_texto(novo_nome, "nome", 100)
                novo_username = limpar_texto(novo_username, "username", 50)
                novo_cargo = limpar_texto(novo_cargo, "cargo", 50)

                if not novo_nome or len(novo_nome.split()) < 2:
                    erros.append("❌ O nome deve conter pelo menos um sobrenome.")
                if not novo_username:
                    erros.append("❌ O nome de usuário é obrigatório.")
                elif novo_username != username_atual and username_em_uso(novo_username):
                    erros.append("❌ Este nome de usuário já está em uso.")
                if not novo_cargo:
                    erros.append("❌ O cargo é obrigatório.")
                if not validar_email(novo_email):
                    erros.append("❌ O email informado é inválido.")
                with conn.cursor() as cursor:
                    cursor.execute("SELECT COUNT(*) FROM FUNCIONARIOS WHERE EMAIL = %s AND ID != %s", (novo_email, id_usuario))
                    if cursor.fetchone()[0] > 0:
                        erros.append("❌ Este email já está em uso por outro funcionário.")

                if erros:
                    for erro in erros:
                        st.error(erro)
                else:
                    @st.dialog("Confirmação")
                    def confirmar_alteracao():
                        st.write("Deseja salvar as alterações?")
                        col1, col2 = st.columns(2)
                        with col1:
                            if st.button("Sim", use_container_width=True):
                                try:
                                    with conn.cursor() as cursor:
                                        usuario_logado = st.session_state["usuario"]["nome"] if "usuario" in st.session_state else "Desconhecido"
                                        cursor.execute("""
                                            UPDATE FUNCIONARIOS 
                                            SET NOME = %s, USERNAME = %s, EMAIL = %s, DTCONTRATACAO = %s, CARGO = %s, ATIVO = %s, ULTIMA_ALTERACAO_POR = %s
                                            WHERE ID = %s
                                        """, (novo_nome, novo_username, novo_email, nova_dt_contratacao, novo_cargo, novo_ativo, usuario_logado, id_usuario))
                                        campos_alterados = {
                                            "NOME": (nome_atual, novo_nome),
                                            "USERNAME": (username_atual, novo_username),
                                            "EMAIL": (email_atual, novo_email),
                                            "CARGO": (cargo_atual, novo_cargo),
                                            "DTCONTRATACAO": (dt_contratacao_atual, nova_dt_contratacao),
                                            "ATIVO": (ativo_atual, novo_ativo)
                                        }
                                        for campo, (anterior, novo) in campos_alterados.items():
                                            if str(anterior) != str(novo):
                                                cursor.execute("""
                                                    INSERT INTO ALTERACOES_FUNCIONARIOS (FUNCIONARIO_ID, ALTERADO_POR, CAMPO_ALTERADO, VALOR_ANTERIOR, VALOR_NOVO)
                                                    VALUES (%s, %s, %s, %s, %s)
                                                """, (id_usuario, usuario_logado, campo, str(anterior), str(novo)))
                                        conn.commit()
                                    st.cache_data.clear()
                                    st.session_state["cache_cleared"] = False
                                    st.session_state.ativo_atual = novo_ativo
                                    st.success("✅ Dados atualizados com sucesso!")
                                    tm.sleep(1)
                                    st.rerun()
                                except Exception as e:
                                    conn.rollback()
                                    st.error(f"❌ Erro ao atualizar os dados: {e}")
                        with col2:
                            if st.button("Não", use_container_width=True):
                                st.info("Alterações canceladas.")

                    confirmar_alteracao()

    if alterar_senha:
        with st.form("form_alterar_senha"):
            nova_senha = st.text_input("Digite a nova senha", type="password", max_chars=100)
            confirmar_senha = st.text_input("Confirme a nova senha", type="password", max_chars=100, key="confirmar_senha_manutencao")

            submit_senha = st.form_submit_button("Alterar Senha", use_container_width=True)

            if submit_senha:
                erros = []
                if not nova_senha or len(nova_senha) < 6:
                    erros.append("❌ A senha deve ter pelo menos 6 caracteres.")
                if nova_senha != confirmar_senha:
                    erros.append("❌ As senhas não coincidem.")

                if erros:
                    for erro in erros:
                        st.error(erro)
                else:
                    @st.dialog("Confirmação")
                    def confirmar_senha():
                        st.write("Deseja alterar a senha?")
                        col1, col2 = st.columns(2)
                        with col1:
                            if st.button("Sim", use_container_width=True):
                                with st.spinner("Alterando senha..."):
                                    try:
                                        with conn.cursor() as cursor:
                                            usuario_logado = st.session_state["usuario"]["nome"] if "usuario" in st.session_state else "Desconhecido"
                                            cursor.execute("UPDATE FUNCIONARIOS SET SENHA = %s, ULTIMA_ALTERACAO_POR = %s WHERE ID = %s",
                                                           (criptografar_senha(nova_senha), usuario_logado, id_usuario))
                                            cursor.execute("""
                                                INSERT INTO ALTERACOES_FUNCIONARIOS (FUNCIONARIO_ID, ALTERADO_POR, CAMPO_ALTERADO, VALOR_ANTERIOR, VALOR_NOVO)
                                                VALUES (%s, %s, %s, %s, %s)
                                            """, (id_usuario, usuario_logado, "SENHA", "****", "****"))
                                            conn.commit()
                                        st.cache_data.clear()
                                        placeholder = st.empty()
                                        placeholder.success("✅ Senha alterada com sucesso!")
                                        tm.sleep(2)
                                        placeholder.empty()
                                        st.rerun()
                                    except Exception as e:
                                        conn.rollback()
                                        st.error(f"❌ Erro ao atualizar a senha: {e}")
                        with col2:
                            if st.button("Não", use_container_width=True):
                                st.info("Alteração de senha cancelada.")

                    confirmar_senha()


#####################################
#Tela para gerenciar os acessos
#####################################

def tela_gerenciar_permissoes():
    st.markdown("<h1 style='text-align: center;'>Gerenciar Acessos</h1>", unsafe_allow_html=True)

    conn = conexao_persistente
    if not conn:
        st.error("Não foi possível conectar ao banco de dados.")
        return

    usuario = st.session_state.get("usuario", {})
    if not usuario.get("gerenciar_permissoes"):
        st.warning("Você não tem permissão para gerenciar permissões de usuários.")
        return

    try:
        with conn.cursor() as cursor:
            # Listar todos os funcionários ativos com suas permissões
            cursor.execute("""
                SELECT ID, NOME, ADMINISTRADOR, CADASTRO, AGENDAMENTO, EDITA_PONTO, AUDITORIA, GERENCIAR_PERMISSOES
                FROM FUNCIONARIOS
                WHERE ATIVO = TRUE
                ORDER BY NOME
            """)
            funcionarios = cursor.fetchall()

            if not funcionarios:
                st.warning("Nenhum funcionário ativo encontrado.")
                return

            # Criar DataFrame original para comparação posterior
            df_original = pd.DataFrame(funcionarios, columns=[
                "ID", "Nome", "Administrador", "Cadastro", "Agendamento", "Edita Ponto", "Auditoria", "Gerenciar Permissões"
            ])
            df = df_original.copy()  # Cópia para exibição e edição
            df = df.replace({'1': True, '0': False})  # Converter para booleanos

            # Criar lista de nomes para o selectbox, incluindo uma opção vazia
            nomes_funcionarios = [""] + [func[1] for func in funcionarios]  # "" representa "Todos"

            # Selectbox para escolher o funcionário
            usuario_selecionado = st.selectbox(
                "Selecione um funcionário (deixe em branco para ver todos):",
                options=nomes_funcionarios,
                index=0  # Começa com a opção vazia selecionada
            )

            # Filtrar o DataFrame se um usuário for selecionado
            if usuario_selecionado:
                df = df[df["Nome"] == usuario_selecionado]

            # Configuração do data_editor
            edited_df = st.data_editor(
                df,
                column_config={
                    "ID": None,
                    #"ID": st.column_config.NumberColumn("ID", disabled=True),
                    "Nome": st.column_config.TextColumn("Funcionário", disabled=True),
                    "Administrador": st.column_config.CheckboxColumn("Admin"),
                    "Cadastro": st.column_config.CheckboxColumn("Cadastro"),
                    "Agendamento": st.column_config.CheckboxColumn("Agenda"),
                    "Edita Ponto": st.column_config.CheckboxColumn("Editar Ponto"),
                    "Auditoria": st.column_config.CheckboxColumn("Auditoria"),
                    "Gerenciar Permissões": st.column_config.CheckboxColumn("Acessos"),
                },
                hide_index=True,
                use_container_width=True,
            )

            # Botão para salvar alterações
            if st.button("Salvar Alterações", use_container_width=True):
                usuario_logado = usuario["nome"]
                alteracoes_realizadas = False  # Flag para verificar se houve alterações

                for _, row in edited_df.iterrows():
                    # Encontrar a linha correspondente no df_original usando o ID
                    original_row = df_original[df_original["ID"] == row["ID"]].iloc[0]
                    
                    # Verificar se houve alterações
                    permissoes_alteradas = {
                        "Administrador": row["Administrador"] != (original_row["Administrador"] == '1'),
                        "Cadastro": row["Cadastro"] != (original_row["Cadastro"] == '1'),
                        "Agendamento": row["Agendamento"] != (original_row["Agendamento"] == '1'),
                        "Edita Ponto": row["Edita Ponto"] != (original_row["Edita Ponto"] == '1'),
                        "Auditoria": row["Auditoria"] != (original_row["Auditoria"] == '1'),
                        "Gerenciar Permissões": row["Gerenciar Permissões"] != (original_row["Gerenciar Permissões"] == '1')
                    }

                    if any(permissoes_alteradas.values()):  # Se houve alguma alteração
                        alteracoes_realizadas = True
                        cursor.execute("""
                            UPDATE FUNCIONARIOS
                            SET ADMINISTRADOR = %s, CADASTRO = %s, AGENDAMENTO = %s, EDITA_PONTO = %s, AUDITORIA = %s, GERENCIAR_PERMISSOES = %s
                            WHERE ID = %s
                        """, (
                            '1' if row["Administrador"] else '0',
                            '1' if row["Cadastro"] else '0',
                            '1' if row["Agendamento"] else '0',
                            '1' if row["Edita Ponto"] else '0',
                            '1' if row["Auditoria"] else '0',
                            '1' if row["Gerenciar Permissões"] else '0',
                            row["ID"]
                        ))

                        # Registrar alterações na tabela ALTERACOES_FUNCIONARIOS
                        for col, alterado in permissoes_alteradas.items():
                            if alterado:
                                cursor.execute("""
                                    INSERT INTO ALTERACOES_FUNCIONARIOS (
                                        FUNCIONARIO_ID, ALTERADO_POR, CAMPO_ALTERADO, VALOR_ANTERIOR, VALOR_NOVO, DATA_ALTERACAO
                                    ) VALUES (%s, %s, %s, %s, %s, %s)
                                """, (
                                    row["ID"],
                                    usuario_logado,
                                    col.upper(),
                                    original_row[col],  # Valor anterior direto do banco
                                    '1' if row[col] else '0',  # Valor novo
                                    datetime.now()
                                ))

                if alteracoes_realizadas:
                    conn.commit()
                    st.success("Acesso alterado! Logon do usuário necessáio!")
                    tm.sleep(2)                    
                    st.rerun()
                else:
                    st.info("Nenhuma alteração detectada para salvar.")
                
    except psycopg2.Error as e:
        conn.rollback()  # Rollback em caso de erro no banco
        st.error(f"Erro no banco de dados: {e}")
    except Exception as e:
        conn.rollback()  # Rollback em caso de erro genérico
        st.error(f"Erro inesperado: {e}")

#Gerar o PDF dos funcionarios cadastrados
#
def gerar_pdf_funcionarios(df):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Lista de Funcionários", ln=True, align="C")
    pdf.ln(5)

    pdf.set_font("Arial", "B", 9)
    pdf.cell(50, 7, "Nome", 1, 0, "C")
    pdf.cell(40, 7, "Cargo", 1, 0, "C")  
    pdf.cell(50, 7, "Email", 1, 0, "C")
    pdf.cell(30, 7, "Data Contratação", 1, 0, "C")
    pdf.cell(20, 7, "Ativo", 1, 1, "C")

    pdf.set_font("Arial", "", 9)
    for _, row in df.iterrows():
        pdf.cell(50, 6, row["Nome"], 1)
        pdf.cell(40, 6, row["Cargo"], 1, 0, "C")  
        pdf.cell(50, 6, row["Email"], 1, 0, "C")
        pdf.cell(30, 6, row["Data Contratação"] if pd.notnull(row["Data Contratação"]) else "Não informado", 1, 0, "C")
        pdf.cell(20, 6, "Sim" if row["Ativo"] else "Não", 1, 1, "C")

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    pdf.output(temp_file.name)
    return temp_file.name

def tela_listar_funcionarios():
    """Tela administrativa para listar todos os funcionários com exportação em XLSX e PDF."""
    if "usuario" not in st.session_state or not st.session_state["usuario"].get("administrador"):
        st.warning("Acesso restrito a administradores.")
        return

    st.markdown("<h1 style='text-align: center;'>Lista de Funcionários</h1>", unsafe_allow_html=True)

    # Checkbox para exibir funcionários inativos
    
    exibir_inativos = st.checkbox("Exibir Funcionários Inativos", value=False)

    conn = conexao_persistente
    if not conn:
        st.error("Erro ao conectar ao banco de dados.")
        return

    try:
        with conn.cursor() as cursor:
            # Ajustar query para incluir ou excluir inativos com base no checkbox
            query = """
                SELECT ATIVO, NOME, CARGO, COALESCE(TO_CHAR(DTCONTRATACAO, 'DD/MM/YYYY'), 'Não informado'), EMAIL
                FROM FUNCIONARIOS
                WHERE ATIVO = %s OR %s = True
                ORDER BY ATIVO, NOME
            """
            cursor.execute("ROLLBACK")
            cursor.execute(query, (not exibir_inativos, exibir_inativos))
            registros = cursor.fetchall()

            if registros:
                df = pd.DataFrame(registros, columns=["Ativo","Nome", "Cargo","Data Contratação", "Email"])
                st.dataframe(df, hide_index=True, use_container_width=True)

                col1, col2 = st.columns(2)
                with col1:
                    # Download em XLSX com colunas ajustadas automaticamente e texto centralizado
                    buffer = BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Funcionarios')
                        workbook = writer.book
                        worksheet = writer.sheets['Funcionarios']
                        
                        # Ajustar largura das colunas automaticamente
                        for column_cells in worksheet.columns:
                            length = max(len(str(cell.value or "")) for cell in column_cells)
                            col_letter = get_column_letter(column_cells[0].column)
                            worksheet.column_dimensions[col_letter].width = length + 2  # Adicionar margem
                            # Centralizar o texto
                            for cell in column_cells:
                                cell.alignment = Alignment(horizontal='center', vertical='center')

                    buffer.seek(0)
                    st.download_button(
                        label="Baixar em Excel",
                        data=buffer.getvalue(),
                        file_name="Lista_Funcionarios.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                with col2:
                    pdf_file = gerar_pdf_funcionarios(df)
                    with open(pdf_file, "rb") as f:
                        st.download_button(
                            label="Baixar em PDF",
                            data=f,
                            file_name="Lista_Funcionarios.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )
                    os.remove(pdf_file)
            else:
                st.info("Nenhum funcionário cadastrado.")
    except Exception as e:
        st.error(f"Erro ao listar funcionários: {e}")


# Função para verificar complexidade da senha
def validar_senha_complexa(senha):
    """Verifica se a senha atende a critérios de complexidade."""
    if not re.search(r"[A-Z]", senha):
        return False, "A senha deve conter pelo menos uma letra maiúscula."
    if not re.search(r"[0-9]", senha):
        return False, "A senha deve conter pelo menos um número."
    if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", senha):
        return False, "A senha deve conter pelo menos um caractere especial."
    return True, ""

# Função para limpar entradas (mantida igual)
def limpar_texto(valor, tipo, max_length):
    """Remove caracteres inválidos e limita o tamanho do campo."""
    if not valor:
        return ""
    if tipo in ["nome", "cargo"]:
        valor = re.sub(r"[^A-Za-zÀ-ÖØ-öø-ÿ\s]", "", valor)
    elif tipo == "username":
        valor = re.sub(r"[^A-Za-z0-9]", "", valor)
    return valor[:max_length]

# Função para validar email (mantida igual)
def validar_email(valor):
    """Valida se o email está no formato correto."""
    if not valor or len(valor) > 100:
        return False
    return bool(re.fullmatch(r"^[\w\.-]+@[\w\.-]+\.\w+$", valor))

# Função para verificar username com cache
@st.cache_data(ttl=300)  # Cache por 5 minutos
def username_em_uso(username):
    """Verifica se o username já está em uso no banco de dados."""
    conn = conexao_persistente
    if not conn:
        return False
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM funcionarios WHERE username = %s", (username,))
            return cursor.fetchone()[0] > 0
    except Exception:
        return False
    
#########################################
#Tela para cadastrar funcionários
##########################################
def cadastrar_funcionarios():
    st.markdown("<h1 style='text-align: center;'>Cadastrar Funcionário</h1>", unsafe_allow_html=True)

    conn = conexao_persistente
    if not conn:
        st.error("Não foi possível conectar ao banco de dados.")
        return

    if "usuario" not in st.session_state or not st.session_state["usuario"].get("cadastro"):
        st.warning("Acesso restrito a usuários com permissão de cadastro.")
        return

    @st.dialog("Confirmar Cadastro")
    def confirmar_cadastro(nome, username, email, cargo, dtcontratacao, senha):
        st.write(f"Você está prestes a cadastrar o funcionário {nome}. Confirme os dados:")
        st.write(f"Nome: {nome}")
        st.write(f"Usuário: {username}")
        st.write(f"E-mail: {email}")
        st.write(f"Cargo: {cargo}")
        st.write(f"Data de Contratação: {dtcontratacao.strftime('%d/%m/%Y')}")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Confirmar", use_container_width=True):
                try:
                    with conn.cursor() as cursor:
                        senha_criptografada = criptografar_senha(senha)
                        cursor.execute("""
                            INSERT INTO FUNCIONARIOS (NOME, USERNAME, EMAIL, CARGO, DTCONTRATACAO, SENHA, ATIVO)
                            VALUES (%s, %s, %s, %s, %s, %s, TRUE)
                        """, (nome, username, email, cargo, dtcontratacao, senha_criptografada))
                        conn.commit()
                        st.session_state["dialog_open"] = False
                        st.rerun()
                except psycopg2.Error as e:
                    conn.rollback()
                    st.error(f"Erro ao cadastrar funcionário: {e}")
        with col2:
            if st.button("Cancelar", use_container_width=True):
                st.session_state["dialog_open"] = False
                st.rerun()

    with st.form(key="form_cadastro"):
        novo_nome = st.text_input("Nome Completo")
        novo_username = st.text_input("Nome de Usuário")
        
        col1, col2 = st.columns(2)
        with col1:
            novo_cargo = st.text_input("Cargo")
        with col2:
            dtcontratacao = st.date_input("Data de Contratação", format="DD/MM/YYYY")
        col3, col4 = st.columns(2)
        with col3:
            senha = st.text_input("Senha", type="password")
        with col4:
            confirmar_senha = st.text_input("Confirme a Senha", type="password") 
        novo_email = st.text_input("E-mail")
        
        submit_button = st.form_submit_button("Cadastrar", use_container_width=True)

        if submit_button:
            erros = []
            novo_nome = limpar_texto(novo_nome, "nome", 100)
            novo_username = limpar_texto(novo_username, "username", 50)
            novo_cargo = limpar_texto(novo_cargo, "cargo", 50)

            if not novo_nome or len(novo_nome.split()) < 2:
                erros.append("❌ O nome deve conter pelo menos um sobrenome.")
            if not novo_username:
                erros.append("❌ O nome de usuário é obrigatório.")
            elif username_em_uso(novo_username):
                erros.append("❌ Este nome de usuário já está em uso.")
            if not novo_cargo:
                erros.append("❌ O cargo é obrigatório.")
            if not validar_email(novo_email):
                erros.append("❌ O e-mail informado é inválido.")
            with conn.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM FUNCIONARIOS WHERE EMAIL = %s", (novo_email,))
                if cursor.fetchone()[0] > 0:
                    erros.append("❌ Este e-mail já está em uso por outro funcionário.")
            if not senha:
                erros.append("❌ A senha é obrigatória.")
            elif len(senha) < 6:
                erros.append("❌ A senha deve ter pelo menos 6 caracteres.")
            elif senha != confirmar_senha:
                erros.append("❌ As senhas não coincidem.")

            if erros:
                for erro in erros:
                    st.error(erro)
            else:
                if "dialog_open" not in st.session_state or not st.session_state["dialog_open"]:
                    st.session_state["dialog_open"] = True
                    confirmar_cadastro(novo_nome, novo_username, novo_email, novo_cargo, dtcontratacao, senha)


#########################################
#Tela para exbir as faltas dos funcionários
##########################################
def tela_admin_faltas():
    st.markdown("<h1 style='text-align: center;'>Faltas Registradas</h1>", unsafe_allow_html=True)

    # Conexão com o banco de dados
    conn = conexao_persistente
    if not conn:
        st.error("Erro na conexão com o banco de dados.")
        return

    with conn.cursor() as cursor:
        cursor.execute("ROLLBACK")
        cursor.execute("""
            SELECT F.FUNCIONARIO_ID, F.DATA, F.JUSTIFICATIVA, F.DOCUMENTO, FUNC.NOME
            FROM FALTAS F
            JOIN FUNCIONARIOS FUNC ON F.FUNCIONARIO_ID = FUNC.ID
            ORDER BY F.DATA DESC
        """)
        registros = cursor.fetchall()

        if registros:
            # Criar dataframe para exibição
            dados = []
            documentos = []  # Lista para armazenar documentos e informações associadas

            for registro in registros:
                id_funcionario, data, justificativa, documento, nome = registro
                possui_documento = "Sim" if documento else "Não"
                dados.append({
                    "X": False,
                    "Data": data.strftime("%d/%m/%Y"),
                    "Funcionário": nome,
                    "Justificativa": justificativa,
                    "Anexo": possui_documento
                })
                documentos.append({
                    "Funcionário": nome,
                    "Data": data.strftime("%d/%m/%Y"),
                    "Anexo": documento
                })

            # Criar dataframe
            df = pd.DataFrame(dados)

            # Filtros em colunas
            col1, col2, col3 = st.columns(3)
            with col1:
                # Filtro de funcionários com múltipla seleção
                funcionarios = df["Funcionário"].unique()
                funcionarios_selecionados = st.multiselect("Selecione os Funcionários", options=funcionarios, default=[])
            with col2:
                # Filtro de data início
                data_inicio = st.date_input("Data Início", value=None, format="DD/MM/YYYY")
            with col3:
                # Filtro de data fim
                data_fim = st.date_input("Data Fim", value=None, format="DD/MM/YYYY")

            # Filtrar dataframe com base nos filtros
            df_filtrado = df.copy()
            if funcionarios_selecionados:
                df_filtrado = df_filtrado[df_filtrado["Funcionário"].isin(funcionarios_selecionados)]
            if data_inicio and data_fim:
                df_filtrado["Data"] = pd.to_datetime(df_filtrado["Data"], format="%d/%m/%Y")
                df_filtrado = df_filtrado[
                    (df_filtrado["Data"] >= pd.to_datetime(data_inicio)) & 
                    (df_filtrado["Data"] <= pd.to_datetime(data_fim))
                ]
                df_filtrado["Data"] = df_filtrado["Data"].dt.strftime("%d/%m/%Y")
            elif data_inicio:
                df_filtrado["Data"] = pd.to_datetime(df_filtrado["Data"], format="%d/%m/%Y")
                df_filtrado = df_filtrado[df_filtrado["Data"] >= pd.to_datetime(data_inicio)]
                df_filtrado["Data"] = df_filtrado["Data"].dt.strftime("%d/%m/%Y")
            elif data_fim:
                df_filtrado["Data"] = pd.to_datetime(df_filtrado["Data"], format="%d/%m/%Y")
                df_filtrado = df_filtrado[df_filtrado["Data"] <= pd.to_datetime(data_fim)]
                df_filtrado["Data"] = df_filtrado["Data"].dt.strftime("%d/%m/%Y")

            # Configurar colunas para DataFrame interativo
            edited_df = st.data_editor(
                df_filtrado,
                column_config={
                    "X": st.column_config.CheckboxColumn(
                        "X",
                        help="Marque para selecionar os documentos que deseja baixar.",
                    ),
                    "Data": "Data",
                    "Funcionário": "Funcionário",
                    "Justificativa": "Justificativa",
                    "Anexo": "Anexo",
                },
                hide_index=True,
                disabled=("Data", "Funcionário", "Justificativa", "Anexo"),
                use_container_width=True,
            )

            # Botões de download em XLSX e PDF
            col4, col5, col6 = st.columns(3)
            with col4:
                # Coletar documentos selecionados para download
                selecionados = edited_df[edited_df["X"] == True]
                if not selecionados.empty:
                    for _, row in selecionados.iterrows():
                        doc_info = next((doc for doc in documentos if doc["Funcionário"] == row["Funcionário"] and doc["Data"] == row["Data"]), None)
                        if doc_info and doc_info["Anexo"]:
                            st.download_button(
                                use_container_width=True,
                                label=f"Baixar Anexo",
                                data=bytes(doc_info["Anexo"]),
                                file_name=f"documento_{row['Funcionário']}_{row['Data']}.pdf",
                                mime="application/pdf",
                            )

            with col5:
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    edited_df.to_excel(writer, index=False, sheet_name='Faltas')
                st.download_button(
                    label="Baixar em Excel",
                    data=buffer.getvalue(),
                    file_name="Faltas_Registradas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            with col6:
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", "B", 12)
                pdf.cell(0, 10, "Faltas Registradas", ln=True, align="C")
                pdf.ln(5)
                pdf.set_font("Arial", "B", 9)
                pdf.cell(20, 7, "Data", 1, 0, "C")
                pdf.cell(50, 7, "Funcionário", 1, 0, "C")
                pdf.cell(90, 7, "Justificativa", 1, 0, "C")
                pdf.cell(30, 7, "Anexo", 1, 1, "C")
                pdf.set_font("Arial", "", 9)
                for _, row in edited_df.iterrows():
                    pdf.cell(20, 6, row["Data"], 1, 0, "C")
                    pdf.cell(50, 6, row["Funcionário"], 1, 0, "C")
                    pdf.cell(90, 6, row["Justificativa"], 1, 0, "C")
                    pdf.cell(30, 6, row["Anexo"], 1, 1, "C")
                pdf_buffer = BytesIO()
                pdf_bytes = pdf.output(dest='S').encode('latin1')  # Gera string e converte para bytes com latin1
                pdf_buffer.write(pdf_bytes)  # Escreve os bytes no BytesIO
                pdf_buffer.seek(0)  # Volta ao início do buffer
                st.download_button(
                    label="Baixar em PDF",
                    data=pdf_buffer.getvalue(),
                    file_name="Faltas_Registradas.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )

        else:
            st.info("Nenhuma falta registrada até o momento.")


def tela_funcionarios_ponto_dia():
    """Tela administrativa para exibir funcionários que registraram ponto em um dia específico."""
    if "usuario" not in st.session_state or not st.session_state["usuario"].get("administrador"):
        st.warning("Acesso restrito a administradores. Faça login como administrador para acessar esta área.")
        return

    st.markdown("<h1 style='text-align: center;'>Funcionários do Dia</h1>", unsafe_allow_html=True)
    st.write("Selecione a data para visualizar os funcionários que registraram ponto:")

    data_selecionada = st.date_input("Data", value=datetime.now().date(), format="DD/MM/YYYY")

    conn = conexao_persistente
    if not conn:
        st.error("Erro ao conectar ao banco de dados.")
        return

    try:
        with conn.cursor() as cursor:
            cursor.execute("ROLLBACK")
            cursor.execute("""
                SELECT DISTINCT f.NOME, f.CARGO, r.CHEGADA, r.SAIDA_ALMOCO, r.RETORNO_ALMOCO, r.SAIDA
                FROM REGISTROS r
                JOIN FUNCIONARIOS f ON r.FUNCIONARIO_ID = f.ID
                WHERE r.DATA = %s
                ORDER BY f.NOME
            """, (data_selecionada,))
            registros = cursor.fetchall()

            if registros:
                df = pd.DataFrame(registros, columns=["Nome", "Cargo", "Chegada", "Saída Almoço", "Retorno Almoço", "Saída"])
                df["Chegada"] = df["Chegada"].apply(formatar_horario)
                df["Saída Almoço"] = df["Saída Almoço"].apply(formatar_horario)
                df["Retorno Almoço"] = df["Retorno Almoço"].apply(formatar_horario)
                df["Saída"] = df["Saída"].apply(formatar_horario)

                st.subheader(f"Funcionários que registraram ponto em {data_selecionada.strftime('%d/%m/%Y')}")
                st.dataframe(df, hide_index=True, use_container_width=True)

                total_funcionarios = len(df)
                st.write(f"**Total de funcionários registrados no dia:** {total_funcionarios}")

                # Botões em colunas
                col1, col2 = st.columns(2)
                with col1:
                    # Baixar em XLSX
                    buffer = BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Funcionarios')
                    st.download_button(
                        label="Baixar Relatório em Excel",
                        data=buffer.getvalue(),
                        file_name=f"Funcionarios_Ponto_{data_selecionada.strftime('%d/%m/%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                with col2:
                    # Baixar em PDF
                    pdf_file = gerar_pdf_funcionarios_dia(df, data_selecionada)
                    with open(pdf_file, "rb") as f:
                        st.download_button(
                            label="Baixar Relatório em PDF",
                            data=f,
                            file_name=f"Funcionarios_Ponto_{data_selecionada.strftime('%d/%m/%Y')}.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )
                    os.remove(pdf_file)

            else:
                st.info(f"Nenhum funcionário registrou ponto em {data_selecionada.strftime('%d/%m/%Y')}.")

            # Funcionários ausentes
            cursor.execute("ROLLBACK")
            cursor.execute("""
                SELECT NOME, CARGO
                FROM FUNCIONARIOS
                WHERE ATIVO = TRUE AND ID NOT IN (
                    SELECT FUNCIONARIO_ID
                    FROM REGISTROS
                    WHERE DATA = %s 
                )
                AND ID NOT IN (
                    SELECT FUNCIONARIO_ID
                    FROM FERIAS
                    WHERE %s BETWEEN DATA_INICIO AND DATA_FIM
                )
                AND ID NOT IN (
                    SELECT FUNCIONARIO_ID
                    FROM FALTAS
                    WHERE DATA = %s
                )
                ORDER BY NOME
            """, (data_selecionada, data_selecionada, data_selecionada))
            ausentes = cursor.fetchall()

            if ausentes:
                df_ausentes = pd.DataFrame(ausentes, columns=["Nome", "Cargo"])
                st.subheader(f"Funcionários ausentes em {data_selecionada.strftime('%d/%m/%Y')}")
                st.dataframe(df_ausentes, hide_index=True, use_container_width=True)
                st.write(f"**Total de funcionários ausentes:** {len(df_ausentes)}")
            else:
                st.info(f"Todos os funcionários registraram ponto ou estão de férias/faltas justificadas em {data_selecionada.strftime('%d/%m/%Y')}.")
    except Exception as e:
        st.error(f"Erro ao buscar registros: {e}")

def gerar_pdf_funcionarios_dia(df, data_selecionada):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"Funcionários que Registraram Ponto - {data_selecionada.strftime('%d/%m/%Y')}", ln=True, align="C")
    pdf.ln(5)

    pdf.set_font("Arial", "B", 9)
    pdf.cell(50, 7, "Nome", 1)
    pdf.cell(40, 7, "Cargo", 1)
    pdf.cell(25, 7, "Chegada", 1, 0, "C")        
    pdf.cell(25, 7, "Saída Almoço", 1, 0, "C")   
    pdf.cell(28, 7, "Retorno Almoço", 1, 0, "C")
    pdf.cell(22, 7, "Saída", 1, 1, "C")

    pdf.set_font("Arial", "", 9)
    for _, row in df.iterrows():
        pdf.cell(50, 6, row["Nome"], 1)
        pdf.cell(40, 6, row["Cargo"], 1)
        pdf.cell(25, 6, row["Chegada"], 1, 0, "C")        
        pdf.cell(25, 6, row["Saída Almoço"], 1, 0, "C")   
        pdf.cell(28, 6, row["Retorno Almoço"], 1, 0, "C") 
        pdf.cell(22, 6, row["Saída"], 1, 1, "C")          

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    pdf.output(temp_file.name)
    return temp_file.name


#Dashboard
def tela_dashboard_admin():
    """Dashboard administrativo com visão geral do sistema de ponto, incluindo férias."""
    if "usuario" not in st.session_state or not st.session_state["usuario"].get("administrador"):
        st.warning("Acesso restrito a administradores. Faça login como administrador para acessar esta área.")
        return

    st.markdown("<h1 style='text-align: center;'>Dashboard Administrativo</h1>", unsafe_allow_html=True)
    st.write("")

    conn = conexao_persistente
    if not conn:
        st.error("Erro ao conectar ao banco de dados.")
        return

    try:
        with conn.cursor() as cursor:
            # Total de funcionários ativos
            cursor.execute("SELECT COUNT(*) FROM FUNCIONARIOS WHERE ATIVO = TRUE")
            total_funcionarios_ativos = cursor.fetchone()[0]

            # Funcionários presentes hoje (ativos que bateram ponto)
            data_atual = datetime.now().date()
            cursor.execute("ROLLBACK")
            cursor.execute("""
                SELECT COUNT(DISTINCT FUNCIONARIO_ID) 
                FROM REGISTROS 
                WHERE DATA = %s 
                AND FUNCIONARIO_ID IN (SELECT ID FROM FUNCIONARIOS WHERE ATIVO = TRUE)
            """, (data_atual,))
            presentes_hoje = cursor.fetchone()[0]

            # Horas extras acumuladas no mês
            primeiro_dia_mes = datetime(data_atual.year, data_atual.month, 1)
            cursor.execute("ROLLBACK")
            cursor.execute("""
                SELECT SUM(EXTRACT(EPOCH FROM HORAEXTRA))
                FROM REGISTROS
                WHERE DATA >= %s AND HORAEXTRA IS NOT NULL
            """, (primeiro_dia_mes,))
            total_segundos = cursor.fetchone()[0] or 0
            horas_extras_mes = total_segundos // 3600
            minutos_extras_mes = (total_segundos % 3600) // 60

            # Faltas pendentes de justificativa
            cursor.execute("ROLLBACK")
            cursor.execute("""
                SELECT COUNT(*)
                FROM FALTAS
                WHERE EXTRACT(MONTH FROM DATA) = EXTRACT(MONTH FROM CURRENT_DATE)
                AND EXTRACT(YEAR FROM DATA) = EXTRACT(YEAR FROM CURRENT_DATE)

            """)
            faltas_pendentes = cursor.fetchone()[0]

            # Funcionários atualmente em férias
            cursor.execute("ROLLBACK")
            cursor.execute("""
                SELECT f.NOME, ferias.DATA_INICIO, ferias.DATA_FIM
                FROM FERIAS ferias
                JOIN FUNCIONARIOS f ON ferias.FUNCIONARIO_ID = f.ID
                WHERE %s BETWEEN ferias.DATA_INICIO AND ferias.DATA_FIM
                ORDER BY f.NOME
            """, (data_atual,))
            ferias_atuais = cursor.fetchall()
            df_ferias_atuais = pd.DataFrame(
                ferias_atuais,
                columns=["Funcionário", "Início", "Fim"]
            )
            df_ferias_atuais["Início"] = pd.to_datetime(df_ferias_atuais["Início"]).dt.strftime("%d/%m/%Y")
            df_ferias_atuais["Fim"] = pd.to_datetime(df_ferias_atuais["Fim"]).dt.strftime("%d/%m/%Y")

            # Funcionários com férias próximas (próximos 30 dias)
            data_fim_proximas = data_atual + timedelta(days=30)
            cursor.execute("ROLLBACK")
            cursor.execute("""
                SELECT f.NOME, ferias.DATA_INICIO, ferias.DATA_FIM
                FROM FERIAS ferias
                JOIN FUNCIONARIOS f ON ferias.FUNCIONARIO_ID = f.ID
                WHERE ferias.DATA_INICIO BETWEEN %s AND %s
                AND ferias.DATA_INICIO > %s
                ORDER BY ferias.DATA_INICIO
            """, (data_atual, data_fim_proximas, data_atual))
            ferias_proximas = cursor.fetchall()
            df_ferias_proximas = pd.DataFrame(
                ferias_proximas,
                columns=["Funcionário", "Início", "Fim"]
            )
            df_ferias_proximas["Início"] = pd.to_datetime(df_ferias_proximas["Início"]).dt.strftime("%d/%m/%Y")
            df_ferias_proximas["Fim"] = pd.to_datetime(df_ferias_proximas["Fim"]).dt.strftime("%d/%m/%Y")

            # Gráfico de férias por mês (próximos 12 meses)
            data_fim_ano = data_atual + timedelta(days=365)
            cursor.execute("ROLLBACK")
            cursor.execute("""
                SELECT DATE_TRUNC('month', ferias.DATA_INICIO) AS Mes, COUNT(*) AS Quantidade
                FROM FERIAS ferias
                WHERE ferias.DATA_INICIO BETWEEN %s AND %s
                GROUP BY DATE_TRUNC('month', ferias.DATA_INICIO)
                ORDER BY Mes
            """, (data_atual, data_fim_ano))
            ferias_por_mes = cursor.fetchall()
            df_ferias_mes = pd.DataFrame(
                ferias_por_mes,
                columns=["Mês", "Quantidade"]
            )
            df_ferias_mes["Mês"] = pd.to_datetime(df_ferias_mes["Mês"]).dt.strftime("%m/%Y")

            # Layout do Dashboard
            # Métricas principais
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Total de Funcionários Ativos", total_funcionarios_ativos)
                st.metric("Presentes Hoje", presentes_hoje,)
            with col2:
                st.metric("Horas Extras (Mês)", f"{horas_extras_mes}h {minutos_extras_mes}m")
                st.metric("Faltas Registradas (Mês)", faltas_pendentes)
            col3, col4 = st.columns(2)
            with col3:
                # Gráfico de pizza: Total de Funcionários Ativos vs Presentes Hoje
                if total_funcionarios_ativos > 0:
                    ausentes_hoje = total_funcionarios_ativos - presentes_hoje
                    pie_data = {
                        "Categoria": ["Presentes", "Ausentes"],
                        "Quantidade": [presentes_hoje, ausentes_hoje]
                    }
                    pie_df = pd.DataFrame(pie_data)
                    fig = {
                        "data": [
                            {
                                "values": pie_df["Quantidade"],
                                "labels": pie_df["Categoria"],
                                "type": "pie",
                                "textinfo": "label+percent",
                                "automargin": True
                            }
                        ],
                        "layout": {
                            "title": "Relação de registros do dia",
                            "height": 400,
                            "width": 300
                        }
                    }
                    st.plotly_chart(fig)
                else:
                    st.info("Nenhum funcionário ativo cadastrado.")

            with col4:
                # Gráfico de presença nos últimos 7 dias (relação presença/total de funcionários ativos)
                data_inicio = data_atual - timedelta(days=6)
                cursor.execute("ROLLBACK")
                cursor.execute("""
                    SELECT 
                        r.DATA, 
                        COUNT(DISTINCT r.FUNCIONARIO_ID) AS Presentes,
                        (SELECT COUNT(ID) FROM FUNCIONARIOS WHERE ATIVO = TRUE) AS TotalFunc
                    FROM REGISTROS r
                    WHERE r.DATA BETWEEN %s AND %s
                    GROUP BY r.DATA
                    HAVING COUNT(DISTINCT r.FUNCIONARIO_ID) > 0
                    ORDER BY r.DATA
                """, (data_inicio, data_atual))
                presencas = cursor.fetchall()

                # Criar DataFrame
                df_presencas = pd.DataFrame(presencas, columns=["Data", "Presentes", "Total de Func"])

                # Converter Data para datetime e formatar para exibição no gráfico
                df_presencas["Data"] = pd.to_datetime(df_presencas["Data"]).dt.strftime("%d/%m")

                # Calcular a porcentagem de presença
                df_presencas["Porcentagem Presença"] = (df_presencas["Presentes"] / df_presencas["Total de Func"] * 100).round(2)

                # Criar o gráfico
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=df_presencas["Data"],
                    y=df_presencas["Porcentagem Presença"],
                    name="Porcentagem de Presença",
                    marker_color='#00cc96',
                    text=[f"{p}%" for p in df_presencas["Porcentagem Presença"]],
                    textposition="auto"
                ))
                fig.update_layout(
                    title="Porcentagem de Presença por Dia",
                    xaxis_title="Data",
                    yaxis_title="Porcentagem de Presença (%)",
                    yaxis_range=[0, 100],
                    bargap=0.2,
                    height=400,
                    width=600
                )

                st.plotly_chart(fig)


            # Seção de Férias
            st.subheader("Funcionários em Férias Hoje")
            if not df_ferias_atuais.empty:
                st.dataframe(df_ferias_atuais, hide_index=True, use_container_width=True)
                st.write(f"**Total em férias hoje:** {len(df_ferias_atuais)}")
            else:
                st.info("Nenhum funcionário em férias hoje.")

            st.subheader("Férias Próximas (30 Dias)")
            if not df_ferias_proximas.empty:
                st.dataframe(df_ferias_proximas, hide_index=True, use_container_width=True)
                st.write(f"**Total de férias próximas:** {len(df_ferias_proximas)}")
            else:
                st.info("Nenhuma férias agendada para os próximos 30 dias.")

            # Gráfico de Férias por Mês
            st.subheader("Distribuição de Férias (Próximos 12 Meses)")
            if not df_ferias_mes.empty:
                st.bar_chart(df_ferias_mes.set_index("Mês"))
            else:
                st.info("Nenhum dado de férias disponível para os próximos 12 meses.")

            # Top 5 funcionários com mais horas extras
            cursor.execute("ROLLBACK")
            cursor.execute("""
                SELECT f.NOME, SUM(EXTRACT(EPOCH FROM r.HORAEXTRA)) as Segundos
                FROM REGISTROS r
                JOIN FUNCIONARIOS f ON r.FUNCIONARIO_ID = f.ID
                WHERE r.HORAEXTRA IS NOT NULL AND r.DATA >= %s
                GROUP BY f.NOME
                ORDER BY Segundos DESC
                LIMIT 5
            """, (primeiro_dia_mes,))
            top_horas = cursor.fetchall()
            df_top_horas = pd.DataFrame(top_horas, columns=["Nome", "Segundos"])
            df_top_horas["Horas"] = df_top_horas["Segundos"].apply(lambda x: f"{x // 3600}h {(x % 3600) // 60}m")
            st.subheader("Top 5 - Horas Extras (Mês)")
            st.dataframe(df_top_horas[["Nome", "Horas"]], hide_index=True, use_container_width=True)

    except Exception as e:
        st.error(f"Erro ao carregar o dashboard: {e}")
#falta do usuario logado

def tela_usuario_faltas():
    usuario = st.session_state["usuario"]
    st.markdown(f"<h1 style='text-align: center;'>Minhas Faltas</h1>", unsafe_allow_html=True)

    # Conexão com o banco de dados
    conn = conexao_persistente
    if conn:
        cursor = conn.cursor()
        cursor.execute("ROLLBACK")
        cursor.execute("""
            SELECT DATA, JUSTIFICATIVA, DOCUMENTO
            FROM FALTAS
            WHERE FUNCIONARIO_ID = %s
            ORDER BY DATA DESC
        """, (usuario["id"],))
        registros = cursor.fetchall()

        if registros:
            # Criar dataframe para exibição
            dados = []
            documentos = []  # Lista para armazenar documentos e informações associadas

            for registro in registros:
                data, justificativa, documento = registro
                possui_documento = "Sim" if documento else "Não"
                dados.append({
                    "X": False,
                    "Data": data.strftime("%d/%m/%Y"),
                    "Justificativa": justificativa,
                    "Anexo": possui_documento
                })

                # Adicionar documento para referência
                documentos.append({
                    "Data": data.strftime("%d/%m/%Y"),
                    "Anexo": documento
                })

            # Criar dataframe
            df = pd.DataFrame(dados)

            # Configurar colunas para DataFrame interativo
            edited_df = st.data_editor(
                df,
                column_config={
                    "X": st.column_config.CheckboxColumn(
                        "X",
                        help="Marque para selecionar os documentos que deseja baixar.",
                    ),
                    "Data": "Data",
                    "Justificativa": "Justificativa",
                    "Anexo": "Anexo",
                },
                hide_index=True,disabled=("Data", "Justificativa","Anexo"),
                use_container_width=True,
            )

            # Coletar documentos selecionados para download
            selecionados = edited_df[edited_df["X"] == True]
            if not selecionados.empty:
                for _, row in selecionados.iterrows():
                    doc_info = next((doc for doc in documentos if doc["Data"] == row["Data"]), None)
                    if doc_info and doc_info["Anexo"]:
                        st.download_button(
                            label="Baixar Documento",
                            data=bytes(doc_info["Anexo"]), 
                            file_name=f"documento_{row['Data']}.pdf",
                            mime="application/pdf",
                        )
            else:
                ''#st.warning("Nenhum documento foi selecionado para download.")
        else:
            st.info("Você não possui faltas registradas!")
    else:
        st.error("Não foi possível conectar ao banco de dados.")


############################################
#Alterações feitas no cadastro do usuario
############################################


def tela_auditoria_cadastros():
    st.markdown("<h1 style='text-align: center;'>Auditoria no Cadastro</h1>", unsafe_allow_html=True)
    conn = conexao_persistente
    if not conn:
        st.error("Erro ao conectar ao banco de dados.")
        return

    try:
        with conn.cursor() as cursor:
            # Filtros
            col1, col2, col3 = st.columns(3)
            with col1:
                usuario_alterado = st.text_input("Usuário Alterado (Nome)", value="")
            with col2:
                data_inicio = st.date_input("Data de Início", value=None, format="DD/MM/YYYY")
            with col3:
                data_fim = st.date_input("Data de Fim", value=None, format="DD/MM/YYYY")
            
            palavra_chave = st.text_input("Palavra-chave no Campo/Valor Alterado", value="")

            # Construir query com filtros
            query = """
                SELECT f.NOME AS "Funcionário", a.ALTERADO_POR, a.CAMPO_ALTERADO , a.VALOR_ANTERIOR, a.VALOR_NOVO, a.DATA_ALTERACAO
                FROM ALTERACOES_FUNCIONARIOS a
                JOIN FUNCIONARIOS f ON a.FUNCIONARIO_ID = f.ID
                WHERE 1=1
            """
            params = []
            if usuario_alterado:
                query += " AND f.NOME ILIKE %s"
                params.append(f"%{usuario_alterado}%")
            if data_inicio:
                query += " AND a.DATA_ALTERACAO >= %s"
                params.append(data_inicio)
            if data_fim:
                query += " AND a.DATA_ALTERACAO <= %s"
                params.append(data_fim)
            if palavra_chave:
                query += " AND (a.VALOR_ANTERIOR ILIKE %s OR a.VALOR_NOVO ILIKE %s OR a.CAMPO_ALTERADO ILIKE %s)"
                params.extend([f"%{palavra_chave}%", f"%{palavra_chave}%",f"%{palavra_chave}%"])
            query += " ORDER BY a.DATA_ALTERACAO DESC"

            cursor.execute(query, params)
            registros = cursor.fetchall()

            if registros:
                df = pd.DataFrame(registros, columns=["Funcionário", "Alterado Por", "Campo Alterado", "Valor Anterior", "Valor Novo", "Data Alteração"])
                df["Data Alteração"] = pd.to_datetime(df["Data Alteração"]).dt.strftime("%d/%m/%Y %H:%M:%S")
                st.dataframe(df, use_container_width=True, hide_index=True)

                # Botões de download
                col1, col2 = st.columns(2)
                with col1:
                    buffer = BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Auditoria_Cadastros')
                        workbook = writer.book
                        worksheet = writer.sheets['Auditoria_Cadastros']
                        for column_cells in worksheet.columns:
                            length = max(len(str(cell.value or "")) for cell in column_cells)
                            col_letter = get_column_letter(column_cells[0].column)
                            worksheet.column_dimensions[col_letter].width = length + 2
                            for cell in column_cells:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                    buffer.seek(0)
                    st.download_button(
                        label="Baixar em XLSX",
                        data=buffer,
                        file_name="Auditoria_Cadastros.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                with col2:
                    # Criar objeto PDF em modo paisagem
                    pdf = FPDF(orientation="L")
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf.set_top_margin(5)
                    pdf.add_page()
                    pdf.set_font("Arial", "B", 12)

                    # Carregar informações da empresa
                    config_empresa = carregar_configuracoes_empresa()
                    nome_empresa = config_empresa.get("nome", "Nome da Empresa")
                    data_emissao = datetime.now().strftime("%d/%m/%Y")

                    pdf.set_font("Arial", "B", 9)  # Negrito

                    # Nome da empresa à esquerda e data de emissão à direita
                    pdf.cell(140, 10, f"Empresa: {nome_empresa}", ln=False, align="L")  
                    pdf.cell(0, 10, f"Data de Emissão: {data_emissao}", ln=True, align="R")  

                    pdf.ln(5)

                    # Título centralizado
                    pdf.set_font("Arial", "B", 12)
                    pdf.cell(0, 13, "Auditoria no Cadastro de Funcionários", ln=True, align="C")
                    pdf.ln(5)

                    # Criar cabeçalho da tabela
                    pdf.set_font("Arial", "B", 9)
                    pdf.cell(59, 7, "Funcionário", 1, 0, "C")
                    pdf.cell(39, 7, "Alterado Por", 1, 0, "C")
                    pdf.cell(30, 7, "Campo", 1, 0, "C")
                    pdf.cell(61, 7, "Valor Anterior", 1, 0, "C")
                    pdf.cell(61, 7, "Valor Novo", 1, 0, "C")
                    pdf.cell(32, 7, "Data Alteração", 1, 1, "C")

                    pdf.set_font("Arial", "", 9)

                    # Função para limitar o tamanho do texto
                    def limitar_texto(texto, largura_maxima):
                        """Trunca o texto se ele ultrapassar a largura máxima da célula."""
                        while pdf.get_string_width(texto) > largura_maxima:
                            texto = texto[:-1]  # Remove o último caractere até caber
                        return texto + "" if len(texto) < len(str(row)) else texto
                    # Divide o nome completo em partes e pega apenas o primeiro e o último nome
                    for _, row in df.iterrows():
                        # Pega apenas o primeiro e o último nome do campo "Alterado Por"
                        nome_completo = row["Alterado Por"].split()
                        if len(nome_completo) > 1:
                            nome_formatado =  f"{nome_completo[0]} {nome_completo[1]}"  # Exemplo: "João Pereira"
                        else:
                            nome_formatado = nome_completo[0] 
                    # Gerar as linhas da tabela
                    for _, row in df.iterrows():
                        pdf.cell(59, 6, limitar_texto(str(row["Funcionário"]), 58), 1, 0, "C")
                        pdf.cell(39, 6, nome_formatado, 1, 0, "C")
                        pdf.cell(30, 6, row["Campo Alterado"], 1, 0, "C")
                        pdf.cell(61, 6, limitar_texto(str(row["Valor Anterior"]), 60), 1, 0, "C")
                        pdf.cell(61, 6, limitar_texto(str(row["Valor Novo"]), 60), 1, 0, "C")
                        pdf.cell(32, 6, row["Data Alteração"], 1, 1, "C")

                    # Gerar o PDF e permitir download
                    pdf_buffer = BytesIO()
                    pdf_bytes = pdf.output(dest='S').encode('latin1')
                    pdf_buffer.write(pdf_bytes)
                    pdf_buffer.seek(0)

                    st.download_button(
                        label="Baixar em PDF",
                        data=pdf_buffer,
                        file_name="Auditoria_Cadastros.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )

            else:
                st.warning("Nenhum registro de alteração encontrado.")
    except Exception as e:
        st.error(f"Erro ao buscar registros de auditoria: {e}")


############################################
#Alterações feitas na folha de ponto
############################################

def tela_auditoria_pontos():
    st.markdown("<h1 style='text-align: center;'>Auditoria no Registro do Ponto</h1>", unsafe_allow_html=True)

    conn = conexao_persistente
    if not conn:
        st.error("Erro ao conectar ao banco de dados.")
        return

    try:
        with conn.cursor() as cursor:
            # Filtros
            col1, col2, col3 = st.columns(3)
            with col1:
                usuario_alterado = st.text_input("Usuário Alterado (Nome)", value="")
            with col2:
                data_inicio = st.date_input("Data de Início", value=None, format="DD/MM/YYYY")
            with col3:
                data_fim = st.date_input("Data de Fim", value=None, format="DD/MM/YYYY")
            
            palavra_chave = st.text_input("Palavra-chave no Valor Alterado", value="")

            # Construir query com filtros
            query = """
                SELECT f.NOME AS "Funcionário", a.ALTERADO_POR AS "Alterado Por", a.CAMPO_ALTERADO AS "Campo Alterado",
                       a.VALOR_ANTERIOR AS "Valor Anterior", a.VALOR_NOVO AS "Valor Novo", a.DATA_ALTERACAO AS "Data Alteração"
                FROM ALTERACOES_REGISTROS_PONTO a
                JOIN FUNCIONARIOS f ON a.FUNCIONARIO_ID = f.ID
                WHERE 1=1
            """
            params = []
            if usuario_alterado:
                query += " AND f.NOME ILIKE %s"
                params.append(f"%{usuario_alterado}%")
            if data_inicio:
                query += " AND a.DATA_ALTERACAO >= %s"
                params.append(data_inicio)
            if data_fim:
                query += " AND a.DATA_ALTERACAO <= %s"
                params.append(data_fim)
            if palavra_chave:
                query += " AND (CAST(a.VALOR_ANTERIOR AS TEXT) ILIKE %s OR CAST(a.VALOR_NOVO AS TEXT) ILIKE %s)"
                params.extend([f"%{palavra_chave}%", f"%{palavra_chave}%"])
            query += " ORDER BY a.DATA_ALTERACAO DESC"

            cursor.execute(query, params)
            registros = cursor.fetchall()

            if registros:
                df = pd.DataFrame(registros, columns=["Funcionário", "Alterado Por", "Campo Alterado", "Valor Anterior", "Valor Novo", "Data Alteração"])
                df["Data Alteração"] = pd.to_datetime(df["Data Alteração"]).dt.strftime("%d/%m/%Y %H:%M:%S")
                df["Valor Anterior"] = df["Valor Anterior"].apply(lambda x: x.strftime("%H:%M:%S") if x else "")
                df["Valor Novo"] = df["Valor Novo"].apply(lambda x: x.strftime("%H:%M:%S") if x else "")
                st.dataframe(df, use_container_width=True, hide_index=True)

                # Botões de download
                col1, col2 = st.columns(2)
                with col1:
                    buffer = BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Auditoria_Pontos')
                        workbook = writer.book
                        worksheet = writer.sheets['Auditoria_Pontos']
                        for column_cells in worksheet.columns:
                            length = max(len(str(cell.value or "")) for cell in column_cells)
                            col_letter = get_column_letter(column_cells[0].column)
                            worksheet.column_dimensions[col_letter].width = length + 2
                            for cell in column_cells:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                    buffer.seek(0)
                    st.download_button(
                        label="Baixar em XLSX",
                        data=buffer,
                        file_name="Auditoria_Pontos.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                with col2:
                    pdf = FPDF(orientation="L")  # Modo paisagem
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf.set_top_margin(5)
                    pdf.set_left_margin(33)  # Margem esquerda para centralizar (297 - 231 = 66 / 2 = 33 mm)
                    pdf.set_right_margin(33)  # Margem direita para centralizar
                    pdf.add_page()

                    config_empresa = carregar_configuracoes_empresa()
                    nome_empresa = config_empresa.get("nome", "Nome da Empresa")
                    data_emissao = datetime.now().strftime("%d/%m/%Y")

                    pdf.set_font("Arial", "B", 10)
                    pdf.cell(140, 10, f"Empresa: {nome_empresa}", ln=False, align="L")
                    pdf.cell(0, 10, f"Data de Emissão: {data_emissao}", ln=True, align="R")
                    pdf.ln(5)

                    pdf.set_font("Arial", "B", 12)
                    pdf.cell(0, 10, "Auditoria no Registro do Ponto", ln=True, align="C")
                    pdf.ln(5)

                    pdf.set_font("Arial", "B", 9)
                    pdf.cell(60, 7, "Funcionário", 1, 0, "C")
                    pdf.cell(39, 7, "Alterado Por", 1, 0, "C")
                    pdf.cell(39, 7, "Campo", 1, 0, "C")
                    pdf.cell(30, 7, "Valor Anterior", 1, 0, "C")
                    pdf.cell(30, 7, "Valor Novo", 1, 0, "C")
                    pdf.cell(33, 7, "Data Alteração", 1, 1, "C")

                    pdf.set_font("Arial", "", 9)
                    
                    def limitar_texto(texto, largura_maxima):
                        """Trunca o texto se ele ultrapassar a largura máxima da célula."""
                        while pdf.get_string_width(texto) > largura_maxima:
                            texto = texto[:-1]  # Remove o último caractere até caber
                        return texto + "" if len(texto) < len(str(row)) else texto

                    for _, row in df.iterrows():
                        nome_completo = row["Alterado Por"].split()
                        if len(nome_completo) > 1:
                            nome_formatado = f"{nome_completo[0]} {nome_completo[1]}"
                        else:
                            nome_formatado = nome_completo[0]

                        pdf.cell(60, 6, limitar_texto(str(row["Funcionário"]),59), 1, 0, "C")
                        pdf.cell(39, 6, nome_formatado, 1, 0, "C")
                        pdf.cell(39, 6, row["Campo Alterado"], 1, 0, "C")
                        pdf.cell(30, 6, row["Valor Anterior"], 1, 0, "C")
                        pdf.cell(30, 6, row["Valor Novo"], 1, 0, "C")
                        pdf.cell(33, 6, row["Data Alteração"], 1, 1, "C")
                    
                    pdf_buffer = BytesIO()
                    pdf_bytes = pdf.output(dest='S').encode('latin1')
                    pdf_buffer.write(pdf_bytes)
                    pdf_buffer.seek(0)
                    st.download_button(
                        label="Baixar em PDF",
                        data=pdf_buffer,
                        file_name="Auditoria_Pontos.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
            else:
                st.warning("Nenhum registro de alteração encontrado.")
    except Exception as e:
        st.error(f"Erro ao buscar registros de auditoria: {e}")

# Escopo para acesso ao Google Calendar
SCOPES = ['https://www.googleapis.com/auth/calendar']

# Função para autenticação usando OAuth2
def autenticar_google_calendar():
    credenciais = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            credenciais = pickle.load(token)

    if not credenciais or not credenciais.valid:
        if credenciais and credenciais.expired and credenciais.refresh_token:
            credenciais.refresh(Request())
        else:
            fluxo = InstalledAppFlow.from_client_secrets_file(
                'AutenticaCalendar.json', SCOPES
            )
            credenciais = fluxo.run_local_server(port=0)
        
        with open('token.pickle', 'wb') as token:
            pickle.dump(credenciais, token)
    
    return credenciais
# Função para obter o ID do calendário autenticado
def obter_id_calendario(credenciais):
    servico = build('calendar', 'v3', credentials=credenciais)
    calendario = servico.calendars().get(calendarId='primary').execute()
    return calendario['id']

#
def adicionar_evento_calendario(servico, calendario_id, titulo, inicio, fim, emails_convidados=None):
    """Adiciona um evento ao Google Calendar"""
    try:
        evento = {
            'summary': titulo,  # Título do evento
            'start': {'dateTime': f"{inicio}T00:00:00-03:00", 'timeZone': 'America/Sao_Paulo'},
            'end': {'dateTime': f"{fim}T23:59:59-03:00", 'timeZone': 'America/Sao_Paulo'},
            'attendees': [{'email': email} for email in emails_convidados] if emails_convidados else [],
        }

        # Envia o evento para o Google Calendar
        servico.events().insert(calendarId=calendario_id, body=evento).execute()
        #print("Evento criado com sucesso:", json.dumps(response, indent=2))
    except:
        ''
    

def obter_usuario_logado():
    usuario = st.session_state.get("usuario", {})
    
    if isinstance(usuario, dict):  # Certifique-se de que 'usuario' é um dicionário
        nome_completo = usuario.get("nome", "Desconhecido")  # Ajuste para pegar o campo correto
    else:
        nome_completo = usuario  # Se for uma string diretamente
    
    partes_nome = nome_completo.split()
    return f"{partes_nome[0]} {partes_nome[1]}" if len(partes_nome) >= 2 else nome_completo


############################################
#Agendamento de férias
############################################

def exibir_formulario_ferias():
    creds = autenticar_google_calendar()
    calendario_id = obter_id_calendario(creds)
    servico = build('calendar', 'v3', credentials=creds)
    
    st.markdown("<h1 style='text-align: center;'>Marcar Férias</h1>", unsafe_allow_html=True)
    
    conn = conexao_persistente
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("ROLLBACK")
            cursor.execute("""
                SELECT ID, NOME, EMAIL
                FROM FUNCIONARIOS
                ORDER BY 2
            """)
            funcionarios = cursor.fetchall()
        except Exception as e:
            st.error(f"Erro ao carregar os funcionários: {e}")
            return
    else:
        st.error("Erro na conexão com o banco de dados.")
        return

    if not funcionarios:
        st.error("Nenhum funcionário encontrado no banco de dados.")
        return
    
    if "funcionarios_dict" not in st.session_state:
        st.session_state["funcionarios_dict"] = {
            f"{nome}": {"id": id, "email": email} for id, nome, email in funcionarios
        }
    
    usuario_logado = obter_usuario_logado()  # Captura usuário logado da sessão
    
    # Inicializar o estado, se necessário
    if "funcionario_selecionado" not in st.session_state:
        primeiro_funcionario = list(st.session_state["funcionarios_dict"].keys())[0]
        st.session_state["funcionario_selecionado"] = primeiro_funcionario
        st.session_state["emails_convidados"] = st.session_state["funcionarios_dict"][primeiro_funcionario]["email"] or ""

    # Atualizar os e-mails convidados ao mudar o funcionário
    def update_emails_convidados():
        """Atualiza os e-mails convidados no estado ao mudar o funcionário."""
        funcionario_atual = st.session_state["funcionario_selecionado"]
        if funcionario_atual in st.session_state["funcionarios_dict"]:
            st.session_state["emails_convidados"] = st.session_state["funcionarios_dict"][funcionario_atual]["email"] or ""

    funcionario_selecionado = st.selectbox(
        "Selecione o Funcionário",
        options=list(st.session_state["funcionarios_dict"].keys()),
        index=list(st.session_state["funcionarios_dict"].keys()).index(st.session_state["funcionario_selecionado"]),
        key="funcionario_selecionado",
        on_change=update_emails_convidados
    )
    
    with st.form("formulario_ferias"):
        col1, col2 = st.columns(2)
        with col1:
            data_inicio = st.date_input("Data de Início", format="DD/MM/YYYY")
        with col2:
            data_fim = st.date_input("Data de Fim", format="DD/MM/YYYY")
        
        ano_referencia = f"{data_inicio.year-1}/{data_inicio.year}"
        
        col3, col4 = st.columns(2)
        with col3:
            st.text_input("Ano de Referência", ano_referencia)
        with col4:
            st.text_input("Responsável pela Autorização", usuario_logado, disabled=True)
        
        emails_convidados = st.text_area(
            "E-mails:",
            placeholder="exemplo1@dominio.com, exemplo2@dominio.com",
            key="emails_convidados",
        )
        
        botao_submeter = st.form_submit_button("Marcar Férias", use_container_width=True)
        
        if botao_submeter:
            if data_inicio > data_fim:
                st.error("A data de início não pode ser maior que a data de fim.")
            elif not funcionario_selecionado:
                st.error("O funcionário é obrigatório.")
            else:
                funcionario_id = st.session_state["funcionarios_dict"][funcionario_selecionado]["id"]
                titulo = f"Férias: {funcionario_selecionado.split(' ')[0]}"
                emails_formatados = [email.strip() for email in emails_convidados.split(",") if email.strip()]
                registrado_em = datetime.now().strftime('%Y-%m-%d')
                
                adicionar_evento_calendario(
                    servico,
                    calendario_id,
                    titulo,
                    data_inicio.strftime('%Y-%m-%d'),
                    (data_fim + timedelta(days=1)).strftime('%Y-%m-%d'),
                    emails_formatados,
                )
                
                if conn:
                    try:
                        cursor.execute("""
                            INSERT INTO FERIAS (FUNCIONARIO_ID, DATA_INICIO, DATA_FIM, ANO_REFERENCIA, RESPONSAVEL_AUTORIZACAO, EMAILS_ENVOLVIDOS,REGISTRADO_EM)
                            VALUES (%s, %s, %s, %s, %s, %s,%s)
                        """, (funcionario_id, data_inicio, data_fim, ano_referencia, usuario_logado, ','.join(emails_formatados),registrado_em))

                        conn.commit()
                        st.success(f"Férias de {funcionario_selecionado.split(' ')[0]} adicionadas ao calendário e salvas no banco de dados!")
                    except Exception as e:
                        st.error(f"Erro ao salvar no banco de dados: {e}")





def minhas_ferias_marcadas():
    usuario = st.session_state["usuario"]
    st.markdown("<h1 style='text-align: center;'>Minhas Férias</h1>", unsafe_allow_html=True)

    # Obter a conexão com o banco de dados
    conn = conexao_persistente
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("ROLLBACK")
            # Buscar todas as férias marcadas
            cursor.execute("""
                SELECT 
                    f.NOME AS Funcionario,
                    ferias.DATA_INICIO AS DataInicio,
                    ferias.DATA_FIM AS DataFim,
                    ferias.REGISTRADO_EM AS RegistradoEm,
                    ferias.ano_referencia
                FROM FERIAS 
                JOIN FUNCIONARIOS f ON ferias.FUNCIONARIO_ID = f.ID
                WHERE FUNCIONARIO_ID = %s
                ORDER BY ferias.DATA_INICIO DESC
            """,(usuario["id"],))
            registros = cursor.fetchall()

            if registros:
                # Criar o DataFrame para exibir os dados
                df = pd.DataFrame(
                    registros,
                    columns=["Funcionário", "Data de Início", "Data Final","Aprovado em","Referência"]
                )

                # Formatando as colunas de data para exibição
                df["Data de Início"] = pd.to_datetime(df["Data de Início"]).dt.strftime("%d/%m/%Y")
                df["Data Final"] = pd.to_datetime(df["Data Final"]).dt.strftime("%d/%m/%Y")
                df["Aprovado em"] = pd.to_datetime(df["Aprovado em"]).dt.strftime("%d/%m/%Y")

                # Exibir o DataFrame no Streamlit
                st.data_editor(df,disabled=True,hide_index=True,use_container_width=True )

            else:
                st.warning("Nenhuma férias marcada no momento.")
        except Exception as e:
            st.error(f"Erro ao buscar férias marcadas: {e}")
        finally:
            ''
    else:
        st.error("Erro na conexão com o banco de dados.")



#Listar férias de todos os funcionários 


def ferias_marcadas():
    st.markdown("<h1 style='text-align: center;'>Férias Agendadas</h1>", unsafe_allow_html=True)

    # Obter a conexão com o banco de dados
    conn = conexao_persistente
    if conn:
        try:
            cursor = conexao_persistente.cursor()
            # Buscar todos os nomes dos funcionários para o filtro
            cursor.execute("SELECT NOME FROM FUNCIONARIOS ORDER BY NOME")
            todos_funcionarios = [row[0] for row in cursor.fetchall()]

            col1, col2, col3 = st.columns(3)
            with col1:
                funcionarios_selecionados = st.multiselect(
                    "Selecione os Funcionários:", options=todos_funcionarios, default=[]
                )
            with col2:
                data_inicio_filtro = st.date_input("Data de Início", value=None, format="DD/MM/YYYY")
            with col3:
                data_fim_filtro = st.date_input("Data de Fim", value=None, format="DD/MM/YYYY")

            # Construir a cláusula WHERE com base na seleção
            filtros = []
            if funcionarios_selecionados:
                filtro_funcionarios = "(" + ", ".join(f"'{nome}'" for nome in funcionarios_selecionados) + ")"
                filtros.append(f"f.NOME IN {filtro_funcionarios}")
            
            if data_inicio_filtro:
                filtros.append(f"ferias.DATA_INICIO >= '{data_inicio_filtro}'")
            
            if data_fim_filtro:
                filtros.append(f"ferias.DATA_FIM <= '{data_fim_filtro}'")
            
            where_clause = " WHERE " + " AND ".join(filtros) if filtros else ""
            
            query = f"""
                SELECT 
                    f.NOME AS Funcionario,
                    ferias.DATA_INICIO,
                    ferias.DATA_FIM,
                    ferias.REGISTRADO_EM,
                    ferias.ano_referencia,
                    ferias.responsavel_autorizacao
                FROM FERIAS
                JOIN FUNCIONARIOS f ON ferias.FUNCIONARIO_ID = f.ID
                {where_clause}
                ORDER BY 1
            """

            # Executar a query e buscar os registros
            cursor.execute(query)
            registros = cursor.fetchall()

            if registros:
                # Criar o DataFrame para exibir os dados
                df = pd.DataFrame(
                    registros,
                    columns=["Funcionário", "Data de Início", "Data Final", "Aprovado em", "Referência", "Autorizado Por"]
                )

                # Formatando as colunas de data para exibição
                df["Data de Início"] = pd.to_datetime(df["Data de Início"]).dt.strftime("%d/%m/%Y")
                df["Data Final"] = pd.to_datetime(df["Data Final"]).dt.strftime("%d/%m/%Y")
                df["Aprovado em"] = pd.to_datetime(df["Aprovado em"]).dt.strftime("%d/%m/%Y")

                # Exibir o DataFrame no Streamlit
                st.dataframe(df, hide_index=True, use_container_width=True)

                # Botões de download em colunas
                col4, col5 = st.columns(2)
                
                with col4:
                    # Download em XLSX com colunas ajustadas automaticamente
                    buffer = BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Ferias_Agendadas')
                        workbook = writer.book
                        worksheet = writer.sheets['Ferias_Agendadas']
                        
                        # Ajustar largura das colunas automaticamente
                        for column_cells in worksheet.columns:
                            length = max(len(str(cell.value or "")) for cell in column_cells)
                            col_letter = get_column_letter(column_cells[0].column)
                            worksheet.column_dimensions[col_letter].width = length + 2  # Adicionar margem
                            # Centralizar o texto
                            for cell in column_cells:
                                cell.alignment = Alignment(horizontal='center', vertical='center')

                    buffer.seek(0)
                    st.download_button(
                        label="Baixar em Excel",
                        data=buffer,
                        file_name="Ferias_Agendadas.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                with col5:
                    # Download em PDF
                    pdf = FPDF()
                    pdf.add_page()
                    pdf.set_font("Arial", "B", 12)
                    pdf.cell(0, 10, "Férias Agendadas", ln=True, align="C")
                    pdf.ln(5)
                    pdf.set_font("Arial", "B", 9)
                    pdf.cell(60, 7, "Funcionário", 1, 0, "C")
                    pdf.cell(25, 7, "Data Início", 1, 0, "C")
                    pdf.cell(25, 7, "Data Fim", 1, 0, "C")
                    pdf.cell(25, 7, "Aprovado em", 1, 0, "C")
                    pdf.cell(25, 7, "Referência", 1, 0, "C")
                    pdf.cell(32, 7, "Autorizado Por", 1, 1, "C")
                    pdf.set_font("Arial", "", 9)
                    for _, row in df.iterrows():
                        pdf.cell(60, 6, row["Funcionário"], 1)
                        pdf.cell(25, 6, row["Data de Início"], 1, 0, "C")
                        pdf.cell(25, 6, row["Data Final"], 1, 0, "C")
                        pdf.cell(25, 6, row["Aprovado em"], 1, 0, "C")
                        pdf.cell(25, 6, row["Referência"], 1, 0, "C")
                        pdf.cell(32, 6, row["Autorizado Por"], 1, 1, "C")
                    pdf_buffer = BytesIO()
                    pdf_bytes = pdf.output(dest='S').encode('latin1')
                    pdf_buffer.write(pdf_bytes)
                    pdf_buffer.seek(0)
                    st.download_button(
                        label="Baixar em PDF",
                        data=pdf_buffer,
                        file_name="Ferias_Agendadas.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )

            else:
                st.warning("Nenhuma férias marcada no momento.")
        except Exception as e:
            st.error(f"Erro ao buscar férias marcadas: {e}")
    else:
        st.error("Erro na conexão com o banco de dados.")


#Adiconar envento no calendario

def add_evento():

    # Autenticar e obter credenciais para o Google Calendar
    creds = autenticar_google_calendar()
    calendario_id = obter_id_calendario(creds)
    servico = build('calendar', 'v3', credentials=creds)

    st.markdown("<h1 style='text-align: center;'>Agendamento</h1>", unsafe_allow_html=True)

    # Obter lista de funcionários com seus e-mails do banco
    conn = conexao_persistente
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("ROLLBACK")
            cursor.execute("""
                SELECT ID, NOME, EMAIL
                FROM FUNCIONARIOS
                ORDER BY 2
            """)
            funcionarios = cursor.fetchall()  # Retorna [(ID, Nome, Email)]
        except Exception as e:
            st.error(f"Erro ao carregar os funcionários: {e}")
            return
    else:
        st.error("Erro na conexão com o banco de dados.")
        return

    if not funcionarios:
        st.error("Nenhum funcionário encontrado no banco de dados.")
        return

    # Criar um dicionário para mapear o nome para ID e e-mail
    funcionarios_dict = {
        nome: {"id": id, "email": email} for id, nome, email in funcionarios
    }

    # Seleção de múltiplos funcionários
    funcionarios_selecionados = st.multiselect(
        "",
        help=None,
        options=list(funcionarios_dict.keys()),
        placeholder="Selecione os Funcionários"
    )

    # Obter os e-mails dos funcionários selecionados
    emails_funcionarios_selecionados = [
        funcionarios_dict[funcionario]["email"]
        for funcionario in funcionarios_selecionados
    ]

    # Formulário para preenchimento dos dados
    with st.form("agendamento"):
        col1, col2 = st.columns(2)
        with col1:
            data = st.date_input("Data", format="DD/MM/YYYY")
        with col2:
            titulo = st.text_input("Título",placeholder="Digite o título da reunião",help="Insira um título descritivo para o agendamento.")

        col3, col4 = st.columns(2)
        with col3:
            hora_inicio = st.time_input("Hora de Início", key="hora_inicio")
        with col4:
            hora_fim = st.time_input("Hora de Fim", key="hora_fim")

        # Preencher o campo de e-mails com os e-mails dos funcionários selecionados
        emails_convidados = st.text_area(
            "E-mails dos Envolvidos (separados por vírgula)",
            value=", ".join(emails_funcionarios_selecionados),  # Usa os e-mails obtidos
            placeholder="exemplo1@dominio.com, exemplo2@dominio.com",
            key="emails_convidados",
        )

        botao_submeter = st.form_submit_button("Agendar",use_container_width=True)
        
    if botao_submeter:
        emails_formatados = [email.strip() for email in emails_convidados.split(",") if email.strip()]

        if not titulo or not emails_formatados:
            st.error("Título e e-mails dos envolvidos são obrigatórios.")
        elif hora_inicio >= hora_fim:
            st.error("A hora de início deve ser menor que a hora de fim.")
        else:
            try:
                # Formatar os horários para o formato ISO 8601
                inicio = f"{data.strftime('%Y-%m-%d')}T{hora_inicio.strftime('%H:%M:%S')}"
                fim = f"{data.strftime('%Y-%m-%d')}T{hora_fim.strftime('%H:%M:%S')}"

                # Adicionar ao Google Calendar
                adicionar_evento_calendario(
                    servico,
                    calendario_id,
                    titulo,
                    inicio,
                    fim,
                    emails_formatados,
                )
                placeholder = st.empty() 
                placeholder.success(f"Evento agendado com sucesso!", icon="✅")
                tm.sleep(2)
                placeholder.empty()
            except Exception as e:
                st.error(f"Erro ao agendar evento: {e}")




# Função para exibição do Google Calendar
def exibir_calendario():
    st.markdown("<h1 style='text-align: center;'>Agenda</h1>", unsafe_allow_html=True)

    calendar_html = '''
    <iframe src="https://calendar.google.com/calendar/embed?height=600&wkst=1&ctz=America%2FSao_Paulo&src=ZGFuaWxvYW5kcmUyN0BnbWFpbC5jb20&color=%234285F4" 
            style="border:solid 1px #777" 
            width="800" 
            height="600" 
            frameborder="0" 
            scrolling="no">
    </iframe>
    '''
    st.markdown(calendar_html, unsafe_allow_html=True)




# Iniciar o aplicativo
if __name__ == "__main__":
    tela_inicial()
